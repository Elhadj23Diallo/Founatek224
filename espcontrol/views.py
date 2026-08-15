""" #Import
from django.shortcuts import render, redirect, (humidity)
            )

            if float(temperature) > 18:
                send_mail(
                    'Alerte : Température élevée',
                    f'Température actuelle : {temperature} °C. Pensez à allumer la clim.',
                    settings.EMAIL_HOST_USER,
                    ['isaacdiallo30@gmail.com'],
                    fail_silently=False,
                )

            return JsonResponse({'temperature': temperature, 'humidity': humidity})
        except ValueError:
            return JsonResponse({'error': 'Valeurs invalides'}, status=400)

    dht_data = DHTData.objects.all().order_by('-created_at')[:10]
    return render(request, 'espcontrol/led_control.html', {
        'dht_data': dht_data,
        'led_status': led.etat
    })

#💡 LED API
@api_permission_required
@api_view(['GET'])
def led_status(request):
    led = LED.objects.first()
    serializer = LEDSerializer(led)
    return Response(serializer.data)

#🌿 Irrigation & Autres interfaces
@api_permission_required
def irrigation_auto(request):
    return render(request, 'espcontrol/irrigation_auto.html')

@api_permission_required
def poubelle_intelligente(request):
    return render(request, 'espcontrol/poubelle_intelligente.html')

@api_permission_required
def control_relais(request):
    return render(request, 'espcontrol/control_relais.html')

@api_permission_required
def get_latest_image(request):
    image = UploadedImage.objects.all().order_by('-created_at').first()
    if image:
        return JsonResponse({'image_url': image.image.url})
    return JsonResponse({'image_url': None})


#🔌 Contrôle Relais via ESP8266

ESP8266_IP = "http://192.168.167.93"

@api_permission_required
def toggle_relais(request, relais_num):
    try:
        url = f"{ESP8266_IP}/relais/{relais_num}/toggle"
        response = requests.get(url)

        if response.status_code == 200:
            return JsonResponse({'status': 'success', 'message': f'Relais {relais_num} contrôlé'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Erreur ESP'}, status=500)

    except requests.exceptions.RequestException as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

#📸 Système de surveillance – upload & affichage image
@api_permission_required
class ImageUploadView(APIView):
    def post(self, request, *args, **kwargs):
        image_data = request.data.get('image', None)
        if not image_data:
            return Response({"error": "No image provided"}, status=400)

        try:
            img_data = base64.b64decode(image_data)
            img = Image.open(BytesIO(img_data))
            img_io = BytesIO()
            img.save(img_io, 'JPEG')
            img_io.seek(0)

            uploaded_image = UploadedImage.objects.create(
                image=ContentFile(img_io.read(), 'received_image.jpg')
            )

            return Response({
                "message": "Image enregistrée avec succès",
                "image_id": uploaded_image.id
            }, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)

@api_permission_required
def systeme_surveillance(request):
    images = UploadedImage.objects.all()
    return render(request, 'espcontrol/surveillance.html', {'images': images})

@api_permission_required
def get_latest_image(request):
    image = UploadedImage.objects.all().order_by('-created_at').first()
    if image:
        return JsonResponse({'image_url': image.image.url})
    return JsonResponse({'image_url': None})


#Vue pour le capteur de gaz
from rest_framework import generics, status
from rest_framework.response import Response
from .models import GasData
from .serializers import GasDataSerializer

API_SECRET_KEY = "@Founatek_2025_SECURITY_KEY!"  # clé à vérifier

class GasDataListCreateView(generics.ListCreateAPIView):
    queryset = GasData.objects.all().order_by('-timestamp')
    serializer_class = GasDataSerializer

    def create(self, request, *args, **kwargs):
        # Récupère la clé secrète dans les données
        secret_key = request.data.get("secret_key")

        if secret_key != API_SECRET_KEY:
            return Response({"error": "Clé API invalide ou manquante"}, status=status.HTTP_403_FORBIDDEN)

        # Supprime la clé avant d’enregistrer les données
        mutable_data = request.data.copy()
        mutable_data.pop("secret_key", None)

        serializer = self.get_serializer(data=mutable_data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

#vue pour afficher les données du capteur de gaz

def gas_data_view(request):
    gas_data = GasData.objects.all().order_by('-timestamp')  # par exemple
    return render(request, 'espcontrol/gas_data.html', {'gas_data': gas_data})

#🌱 Données Capteur d’Humidité du Sol (à compléter)

@api_permission_required
def soil_data(request):
    # Lire l'humidité envoyée par le capteur
    humidity = request.GET.get('humidity')

    # Afficher l'humidité reçue dans les logs pour déboguer
    print(f"Received humidity: {humidity}")

    # Vérifier si l'humidité a bien été reçue et si elle peut être convertie en entier
    if humidity is not None:
        try:
            humidity = int(humidity)  # Convertir en entier explicitement
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid humidity value'}, status=400)

        # Sauvegarder les données d'humidité dans la base de données
        soil_data = SoilData(humidity=humidity)
        soil_data.save()

        # Vérifier si l'humidité dépasse un seuil critique (par exemple, si elle est trop faible)
        if humidity < 30:  # Seuil d'humidité critique
            send_mail(
                'Alerte : Humidité du sol faible',
                f'L\'humidité actuelle du sol est de {humidity}%, ce qui est en dessous du seuil critique de 30%. Pensez à arroser vos plantes.',
                settings.EMAIL_HOST_USER,  # Expéditeur (Doit être configuré dans settings.py)
                ['isaacdiallo30@gmail.com'],  # Liste des destinataires
                fail_silently=False,
            )
        print(f"New DHTData created with Temperature: Humidity: {humidity}")
        # Retourner une réponse JSON avec le statut de succès
        return JsonResponse({'status': 'success', 'humidity': humidity})
    else:
        return JsonResponse({'status': 'error', 'message': 'Missing humidity data'}, status=400)

#🔒 API sécurisée d’exemple
@api_permission_required
class CompteurDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response({"message": "Voici les données du capteur."})


@api_permission_required
def display_soil_data(request):
    # Récupérer toutes les données d'humidité
    data = SoilData.objects.all().order_by('-created_at')  # Trier par date décroissante

    # Passer les données au template
    return render(request, 'espcontrol/soil_data.html', {'data': data})



class CompteurDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Exemple de retour de données protégées
        return Response({"message": "Voici les données du capteur."})

@api_permission_required
class ComptageAPIView(APIView):

    def get(self, request):
        comptages = Comptage.objects.all().order_by('-timestamp')[:10]
        serializer = ComptageSerializer(comptages, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ComptageSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

     """













from django.http import JsonResponse
from django.shortcuts import render, redirect
import requests
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import status
from .models import Comptage, DHTData, LED, UploadedImage, SoilData, Device, AppareilData
from .serializers import ComptageSerializer, LEDSerializer, SensorDataSerializer
from django.core.mail import send_mail
from django.conf import settings
from io import BytesIO
import base64
from PIL import Image
from django.core.files.base import ContentFile
from django.contrib.auth.decorators import login_required
from rest_framework.permissions import IsAuthenticated
from .utils import api_permission_required
try:
    from monetisation.decorators import plan_required, require_quota, premium_feature_required
    from monetisation.quota import check_quota
except Exception:
    plan_required = lambda *args, **kwargs: (lambda view: view)
    require_quota = lambda *args, **kwargs: (lambda view: view)
    premium_feature_required = lambda *args, **kwargs: (lambda view: view)
    check_quota = lambda *args, **kwargs: True
# ... tes imports actuels ...
from django.core.paginator import Paginator # Pour faire des pages (1, 2, 3...)
import zipfile # Pour créer le fichier ZIP
import json
import joblib
import pandas as pd
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from pathlib import Path
try:
    from iot.models import Parcours
except Exception:
    Parcours = None
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import AgentAlert
from .models import Device, AppareilData, AgentAlert
import numpy as np


#alerts
from espcontrol.ml.services.air_alerts import evaluate_air_alert_now
from espcontrol.models import AirReading, AgentAlert
from django.http import JsonResponse


def air_alert_now(request, city):
    reading = AirReading.objects.filter(
        source__icontains=city.upper(),
        is_active=True
    ).order_by("-simulated_time").first()

    if not reading:
        return JsonResponse({"alert": None})

    alert = evaluate_air_alert_now(reading)

    return JsonResponse({
        "city": city,
        "alert": alert,
    })


#carte d'afrique en temps réel
from django.http import JsonResponse
from espcontrol.models import AirReading

CITIES = {
    "conakry":  {"lat": 9.64,  "lon": -13.58},
    "dakar":    {"lat": 14.69, "lon": -17.44},
    "abidjan":  {"lat": 5.36,  "lon": -4.01},
    "lagos":    {"lat": 6.52,  "lon": 3.38},
    "nairobi":  {"lat": -1.29, "lon": 36.82},
}

def air_map_data(request):
    data = []

    for city in CITIES.keys():
        reading = AirReading.objects.filter(
            source__icontains=city.upper(),
            is_active=True
        ).order_by("-simulated_time").first()

        if not reading:
            continue

        data.append({
            "city": city,
            "lat": reading.latitude,
            "lon": reading.longitude,
            "pm2p5": reading.pm2p5,
            "pm10": reading.pm10,
            "co": reading.co,
            "no2": reading.no2,
            "confidence": reading.confidence,
            "origin": reading.origin,
            "source": reading.source,
        })

    return JsonResponse({"data": data})




def air_map_view(request):
    return render(request, "espcontrol/air_map.html")


# espcontrol/views.py (ou où est ta vue)

from django.http import JsonResponse
from espcontrol.ml.services.air_fusion import get_latest_air_reading


def latest_air_data(request, city):
    reading = get_latest_air_reading(city)

    if not reading:
        return JsonResponse({"error": "no data"}, status=404)

    return JsonResponse({
        "city": city,
        "origin": reading.origin,        # virtual / real / hybrid
        "confidence": reading.confidence,
        "simulated_time": reading.simulated_time.isoformat(),
        "created_at": reading.created_at.isoformat(),
        "location": {
            "lat": reading.latitude,
            "lon": reading.longitude,
        },
        "data": {
            "pm2p5": reading.pm2p5,
            "pm10": reading.pm10,
            "co": reading.co,
            "no2": reading.no2,
        },
        "source": reading.source,
    })




#Similateur
# espcontrol/views.py

import json
from pathlib import Path

from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

from espcontrol.models import AirSimulatedData, AirReading


# -------------------------
# Helpers
# -------------------------

def _json_error(message: str, status: int = 400, **extra):
    payload = {"error": message}
    if extra:
        payload.update(extra)
    return JsonResponse(payload, status=status)


def _obj_get(obj, key: str, default=None):
    """
    Supporte obj.key (objet) ou obj[key] (dict).
    """
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


def _serialize_airlike(obj):
    """
    Sérialise un objet/dict contenant :
    simulated_time, created_at, latitude, longitude, pm2p5, pm10, co, no2, source
    """
    simulated_time = _obj_get(obj, "simulated_time")
    created_at = _obj_get(obj, "created_at")

    # isoformat si possible
    if hasattr(simulated_time, "isoformat"):
        simulated_time = simulated_time.isoformat()
    if hasattr(created_at, "isoformat"):
        created_at = created_at.isoformat()

    return {
        "simulated_time": simulated_time,
        "created_at": created_at,
        "location": {
            "lat": _obj_get(obj, "latitude"),
            "lon": _obj_get(obj, "longitude"),
        },
        "data": {
            "pm2p5": _obj_get(obj, "pm2p5"),
            "pm10": _obj_get(obj, "pm10"),
            "co": _obj_get(obj, "co"),
            "no2": _obj_get(obj, "no2"),
        },
        "source": _obj_get(obj, "source"),
    }


# -------------------------
# Air simulator endpoints
# -------------------------

def air_sim_tick(request):
    """
    Lance un 'tick' du simulateur (écrit souvent en DB) et renvoie la donnée générée.
    IMPORTANT: import lazy pour éviter crash WSGI si fichiers ML absents.
    """
    try:
        from espcontrol.ml.services.air_simulator import simulator_tick  # lazy import
    except Exception as e:
        return _json_error("simulator service not available", status=503, detail=str(e))

    try:
        obj = simulator_tick()
        return JsonResponse(_serialize_airlike(obj), status=200)
    except FileNotFoundError as e:
        # typiquement data_mlev.nc absent
        return _json_error("simulator data file missing", status=503, detail=str(e))
    except Exception as e:
        return _json_error("simulator tick failed", status=500, detail=str(e))


def air_sim_latest(request):
    """
    Retourne la dernière simulation enregistrée.
    """
    obj = AirSimulatedData.objects.order_by("-created_at").first()
    if not obj:
        return _json_error("no data", status=404)

    return JsonResponse(_serialize_airlike(obj), status=200)


def latest_air_reading(request):
    """
    Retourne la dernière lecture AirReading provenant de la simulation CAMS_SIM.
    """
    r = (
        AirReading.objects.filter(source="CAMS_SIM")
        .order_by("-simulated_time", "-created_at")
        .first()
    )
    if not r:
        return _json_error("no simulated data yet", status=404)

    return JsonResponse(_serialize_airlike(r), status=200)


# -------------------------
# Air ML endpoints
# -------------------------

def air_alerts(request):
    """
    Calcule des alertes à partir des prédictions ML.
    """
    try:
        from espcontrol.ml.services.air_predictor import predict_conakry_next_3h  # lazy
        from espcontrol.ml.services.air_alerts import evaluate_air_quality  # lazy
    except Exception as e:
        return _json_error("air ML services not available", status=503, detail=str(e))

    try:
        prediction_data = predict_conakry_next_3h()
        alerts = evaluate_air_quality(prediction_data["prediction"])
        return JsonResponse(
            {
                "from_time": prediction_data.get("from_time"),
                "to_time": prediction_data.get("to_time"),
                "alerts": alerts,
            },
            status=200,
        )
    except FileNotFoundError as e:
        return _json_error("ML data file missing", status=503, detail=str(e))
    except Exception as e:
        return _json_error("air alerts failed", status=500, detail=str(e))


@csrf_exempt
def air_predict(request):
    """
    API POST: renvoie la prédiction ML (Conakry next 3h).
    """
    if request.method != "POST":
        return _json_error("POST required", status=405)

    try:
        from espcontrol.ml.services.air_predictor import predict_conakry_next_3h  # lazy
    except Exception as e:
        return _json_error("predictor service not available", status=503, detail=str(e))

    try:
        data = predict_conakry_next_3h()
        return JsonResponse(data, status=200)
    except FileNotFoundError as e:
        return _json_error("ML data file missing", status=503, detail=str(e))
    except Exception as e:
        return _json_error("prediction failed", status=500, detail=str(e))


def air_dashboard(request):
    """
    Dashboard HTML pour les prédictions.
    IMPORTANT: lazy import.
    """
    try:
        from espcontrol.ml.services.air_predictor import predict_conakry_next_3h  # lazy
        from espcontrol.ml.services.air_alerts import evaluate_air_quality  # lazy
    except Exception as e:
        # En dashboard, on peut afficher une page “service indisponible”
        return render(
            request,
            "espcontrol/air_dashboard.html",
            {
                "error": "air ML services not available",
                "detail": str(e),
                "image": "/static/air/pm2p5_dashboard.png",
            },
            status=503,
        )

    try:
        data = predict_conakry_next_3h()
        pm25 = (data.get("prediction") or {}).get("pm2p5")

        # Niveau simple
        level = "Inconnu"
        color = "gray"
        if isinstance(pm25, (int, float)):
            if pm25 < 1e-7:
                level, color = "Bon", "green"
            elif pm25 < 1.5e-7:
                level, color = "Moyen", "orange"
            else:
                level, color = "Mauvais", "red"

        alerts = evaluate_air_quality(data.get("prediction") or {})

        context = {
            "data": data,
            "level": level,
            "color": color,
            "alerts": alerts,
            "image": "/static/air/pm2p5_dashboard.png",
        }
        return render(request, "espcontrol/air_quality_dashboard.html", context)

    except FileNotFoundError as e:
        return render(
            request,
            "espcontrol/air_dashboard.html",
            {
                "error": "ML data file missing",
                "detail": str(e),
                "image": "/static/air/pm2p5_dashboard.png",
            },
            status=503,
        )
    except Exception as e:
        return render(
            request,
            "espcontrol/air_dashboard.html",
            {
                "error": "dashboard failed",
                "detail": str(e),
                "image": "/static/air/pm2p5_dashboard.png",
            },
            status=500,
        )


# -------------------------
# Irrigation ML endpoint
# -------------------------

@csrf_exempt
def irrigation_prediction(request):
    """
    API POST : reçoit des données IoT et retourne ON/OFF.
    IMPORTANT: charge le modèle de façon safe (pas au niveau global).
    """
    if request.method != "POST":
        return _json_error("POST only", status=405)

    try:
        import pandas as pd
        import joblib
        from .MAL.features import FEATURES
    except Exception as e:
        return _json_error("irrigation dependencies not available", status=503, detail=str(e))

    try:
        # Body JSON
        try:
            payload = json.loads(request.body or b"{}")
        except Exception:
            return _json_error("invalid JSON body", status=400)

        # Model path
        model_path = Path(__file__).resolve().parent / "MAL" / "irrigation_model.pkl"
        if not model_path.exists():
            return _json_error("model not available on server", status=503, path=str(model_path))

        model = joblib.load(model_path)

        # Construire le sample avec colonnes attendues
        # (sécurise les KeyError)
        required = ["temperature_c", "humidity_air", "rainfall_mm", "hour"]
        missing = [k for k in required if k not in payload]
        if missing:
            return _json_error("missing fields", status=400, missing=missing)

        sample_df = pd.DataFrame(
            [
                {
                    "temperature_c": payload["temperature_c"],
                    "humidity_air": payload["humidity_air"],
                    "rainfall_mm": payload["rainfall_mm"],
                    "hour": payload["hour"],
                }
            ],
            columns=FEATURES,
        )

        prediction = model.predict(sample_df)[0]
        return JsonResponse(
            {"irrigation": "ON" if int(prediction) == 1 else "OFF", "decision": int(prediction)},
            status=200,
        )

    except FileNotFoundError as e:
        return _json_error("model file missing", status=503, detail=str(e))
    except Exception as e:
        # 400 si input invalide, 500 si bug interne.
        return _json_error("prediction error", status=400, detail=str(e))



# Page d'accueil

@login_required
def home(request):
    user = request.user

    is_abonne = user.groups.filter(name="Abonné").exists()
    is_admin = (
        user.is_superuser
        or user.is_staff
        or user.groups.filter(name="Administrateur").exists()
    )

    # 🔹 Parcours publié (ex: premier parcours disponible)
    parcours = Parcours.objects.filter(is_published=True).first()

    return render(request, "espcontrol/home.html", {
        "is_abonne": is_abonne,
        "is_admin": is_admin,
        "parcours": parcours,   # 👈 AJOUT ICI
    })



# API pour les comptages
class ComptageAPIView(APIView):
    def get(self, request):
        comptages = Comptage.objects.filter(user=request.user).order_by('-timestamp')[:10]
        serializer = ComptageSerializer(comptages, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ComptageSerializer(data=request.data)
        if serializer.is_valid():
            comptage = serializer.save(user=request.user)  # Lier le comptage à l'utilisateur
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# API pour contrôler la LED
@api_view(['GET'])
def led_status(request):
    led = LED.objects.filter(user=request.user).first()  # Filtrer par utilisateur
    if not led:
        return Response({'error': 'LED non trouvée'}, status=status.HTTP_404_NOT_FOUND)
    serializer = LEDSerializer(led)
    return Response(serializer.data)

# Contrôle de la LED avec changement d'état
@api_permission_required
@premium_feature_required(required_plans=['basic', 'pro'])
@require_quota('api')
#@plan_required(['basic', 'pro'])
def led_control(request):
    led, created = LED.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        led.etat = not led.etat
        led.save()
        return redirect('dht-data')

    temperature = request.GET.get('temperature')
    humidity = request.GET.get('humidity')
    print(f"Temperature: {temperature}, Humidity: {humidity}")  # Debug ici

    if temperature and humidity:
        try:
            # Ajouter l'utilisateur lors de l'enregistrement
            DHTData.objects.create(
                temperature=float(temperature),
                humidity=float(humidity),
                user=request.user  # Associer l'utilisateur
            )

            if float(temperature) > 18:
                send_mail(
                    'Alerte : Température élevée',
                    f'Température actuelle : {temperature} °C. Pensez à allumer la clim.',
                    settings.EMAIL_HOST_USER,
                    ['isaacdiallo30@gmail.com'],
                    fail_silently=False,
                )

            return JsonResponse({'temperature': temperature, 'humidity': humidity})
        except ValueError:
            return JsonResponse({'error': 'Valeurs invalides'}, status=400)

    dht_data = DHTData.objects.filter(user=request.user).order_by('-created_at')[:10]
    return render(request, 'espcontrol/led_control.html', {
        'dht_data': dht_data,
        'led_status': led.etat
    })


# Upload d'une image
# API Upload Image (ESP32)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.cache import cache
from .models import UploadedImage
from PIL import Image, ImageChops, ImageStat
import os
import io
import zipfile

# 🔥 CORRECTION DÉFINITIVE DES IMPORTS DE DATE
from datetime import datetime, timedelta

# --- CONFIGURATION ---
LIVE_DIR = "/home/Founatek224/Founatek224/media/live"
MOTION_THRESHOLD = 5.0
CACHE_TIMEOUT = 10  # secondes pour le live

from django.core.cache import cache

# --- VUE D'UPLOAD OPTIMISÉE ---
class ImageUploadView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def detect_motion(self, current_img_path, new_img_data):
        try:
            if not os.path.exists(current_img_path): return 0
            img1 = Image.open(current_img_path).convert("L").resize((64, 64))
            img2 = Image.open(io.BytesIO(new_img_data)).convert("L").resize((64, 64))
            diff = ImageChops.difference(img1, img2)
            stat = ImageStat.Stat(diff)
            return stat.mean[0]
        except:
            return 0

    def post(self, request, *args, **kwargs):
        if 'image' not in request.FILES:
            return Response({"error": "No file"}, status=400)

        image_file = request.FILES['image']
        new_image_data = image_file.read()
        camera_id = request.data.get('camera_id', 'camera_salon') or "camera_salon"

        if not os.path.exists(LIVE_DIR):
            try: os.makedirs(LIVE_DIR)
            except: pass

        # 🔥 AJOUT USER DANS LE NOM DU FICHIER
        user_prefix = f"user_{request.user.id}"
        live_file_path = os.path.join(LIVE_DIR, f"{user_prefix}_{camera_id}.jpg")

        # 1. DÉTECTION
        motion_score = self.detect_motion(live_file_path, new_image_data)
        has_motion = motion_score > MOTION_THRESHOLD

        # 2. ARCHIVAGE SUR DISQUE (si mouvement)
        if has_motion:
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{camera_id}_{timestamp}.jpg"

                UploadedImage.objects.create(
                    image=ContentFile(new_image_data, name=filename),
                    user=request.user,
                    camera_id=camera_id,
                    has_motion=True
                )
            except Exception as e:
                print(f"Erreur DB: {e}")

        # 3. MISE À JOUR LIVE
        try:
            cache.set(f"live_{request.user.id}_{camera_id}", new_image_data, CACHE_TIMEOUT)

            with open(live_file_path, 'wb') as f:
                f.write(new_image_data)
        except Exception as e:
            print(f"Erreur live: {e}")

        return Response({"status": "ok"}, status=200)


# --- VUES DASHBOARD ---

@login_required
def systeme_surveillance(request):
    return render(request, "espcontrol/surveillance.html")


@login_required
def get_latest_frames(request):
    data = []
    camera_ids = []

    user_prefix = f"user_{request.user.id}"

    if os.path.exists(LIVE_DIR):
        for filename in os.listdir(LIVE_DIR):
            if filename.endswith(".jpg") and filename.startswith(user_prefix):
                cam_id = filename.replace(f"{user_prefix}_", "").replace(".jpg", "")
                camera_ids.append(cam_id)

    if not camera_ids:
        camera_ids = ["camera_salon"]

    for cam_id in camera_ids:
        is_alert = False
        try:
            recent = datetime.now() - timedelta(seconds=5)
            if UploadedImage.objects.filter(
                user=request.user,
                camera_id=cam_id,
                has_motion=True,
                created_at__gte=recent
            ).exists():
                is_alert = True
        except:
            pass

        data.append({
            "camera_id": cam_id,
            "image_url": f"/stream/{cam_id}/",
            "created_at": "EN DIRECT",
            "has_motion": is_alert
        })

    return JsonResponse({"cameras": data})


# --- VUE LIVE OPTIMISÉE ---
@login_required
def get_live_image_content(request, camera_id):
    cache_key = f"live_{request.user.id}_{camera_id}"

    image_data = cache.get(cache_key)
    if image_data:
        return HttpResponse(image_data, content_type="image/jpeg")

    user_prefix = f"user_{request.user.id}"
    file_path = os.path.join(LIVE_DIR, f"{user_prefix}_{camera_id}.jpg")

    try:
        with open(file_path, "rb") as f:
            image_data = f.read()
            cache.set(cache_key, image_data, CACHE_TIMEOUT)
            return HttpResponse(image_data, content_type="image/jpeg")
    except FileNotFoundError:
        return HttpResponse(status=404)


# --- VUES ARCHIVE / DOWNLOAD ---

@login_required
def archive_gallery(request):
    from django.core.paginator import Paginator
    image_list = UploadedImage.objects.filter(
        user=request.user,
        has_motion=True
    ).order_by('-created_at')

    paginator = Paginator(image_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "espcontrol/archive.html", {"page_obj": page_obj})


@login_required
def download_zip(request):
    images = UploadedImage.objects.filter(
        user=request.user,
        has_motion=True
    ).order_by('-created_at')[:50]

    if not images:
        return HttpResponse("Aucune image.", status=404)

    response = HttpResponse(content_type='application/zip')

    try:
        now_str = datetime.now().strftime('%Y%m%d_%H%M')
        zip_name = f"Preuves_Founatek_{now_str}.zip"
    except:
        zip_name = "Preuves_Founatek.zip"

    response['Content-Disposition'] = f'attachment; filename={zip_name}'

    with zipfile.ZipFile(response, 'w') as zip_file:
        for img in images:
            if img.image:
                try:
                    zip_file.write(img.image.path, arcname=os.path.basename(img.image.path))
                except:
                    pass

    return response


from django.http import JsonResponse

@login_required
def get_latest_image(request, camera_id):
    image = (
        UploadedImage.objects
        .filter(user=request.user, camera_id=camera_id)
        .order_by('-created_at')
        .first()
    )

    if not image:
        return JsonResponse({"image_url": None})

    return JsonResponse({
        "image_url": image.image.url
    })

# Données Humidité du Sol
from django.http import JsonResponse
from .models import SoilData
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def soil_data(request):
    if request.method == 'POST':
        humidity = request.data.get('humidity')

        if humidity is not None:
            try:
                humidity = int(humidity)
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Invalid humidity value'}, status=400)

            # Sauvegarder l'humidité avec l'utilisateur
            soil_data = SoilData(humidity=humidity, user=request.user)
            soil_data.save()

            if humidity < 30:
                send_mail(
                    'Alerte : Humidité du sol faible',
                    f'L\'humidité actuelle du sol est de {humidity}%, ce qui est en dessous du seuil critique de 30%. Pensez à arroser vos plantes.',
                    settings.EMAIL_HOST_USER,
                    ['isaacdiallo30@gmail.com'],
                    fail_silently=False,
                )

            return JsonResponse({'status': 'success', 'humidity': humidity})
        else:
            return JsonResponse({'status': 'error', 'message': 'Missing humidity data'}, status=400)

    elif request.method == 'GET':
        # Récupérer toutes les données d'humidité de l'utilisateur
        user_soil_data = SoilData.objects.filter(user=request.user)
        data = [{'humidity': item.humidity, 'created_at': item.created_at} for item in user_soil_data]

        return JsonResponse({'status': 'success', 'data': data})


#🌿 Irrigation & Autres interfaces
def irrigation_auto(request):
    return render(request, 'espcontrol/irrigation_auto.html')

def poubelle_intelligente(request):
    return render(request, 'espcontrol/poubelle_intelligente.html')

class CompteurDataAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Exemple de retour de données protégées
        return Response({"message": "Voici les données du capteur."})


# views.py
from django.http import JsonResponse
from .models import Relais
from django.views.decorators.csrf import csrf_exempt
import json

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator

@csrf_exempt
@login_required
def get_relais_etat(request, num):
    try:
        relais = Relais.objects.get(num=num, user=request.user)
        return JsonResponse({"num": relais.num, "etat": "on" if relais.etat else "off"})
    except Relais.DoesNotExist:
        return JsonResponse({"error": "Relais non trouvé"}, status=404)


@csrf_exempt
@login_required
def set_relais_etat(request, num):
    try:
        relais = Relais.objects.get(num=num, user=request.user)
    except Relais.DoesNotExist:
        return JsonResponse({"error": "Relais non trouvé"}, status=404)

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            relais.etat = (data.get("state") == "on")
            relais.save()
            return JsonResponse({
                "num": relais.num,
                "etat": "on" if relais.etat else "off"
            })
        except (KeyError, json.JSONDecodeError):
            return JsonResponse({"error": "Données JSON invalides"}, status=400)

    return JsonResponse({"error": "Méthode non autorisée"}, status=405)



from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import Relais  # Assure-toi que ce modèle est bien importé

@csrf_exempt
@login_required
def get_all_relais_etats(request):
    relais_list = Relais.objects.filter(user=request.user)
    data = [{"relais": r.num, "etat": "on" if r.etat else "off"} for r in relais_list]
    return JsonResponse(data, safe=False)



from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Relais
from django.core.mail import send_mail
from django.conf import settings

# Vue pour contrôler les relais
def relais_control(request):
    # On récupère tous les relais pour l'utilisateur connecté
    relais = Relais.objects.filter(user=request.user)

    # Vérifie si la requête est de type POST (pour activer/désactiver les relais)
    if request.method == 'POST':
        relais_num = request.POST.get('relais_num')  # Numéro du relais
        action = request.POST.get('action')  # Action à réaliser : 'on' ou 'off'

        try:
            relais_obj = relais.get(num=relais_num)
            if action == 'on':
                relais_obj.etat = True
            elif action == 'off':
                relais_obj.etat = False
            else:
                return JsonResponse({'error': 'Action non valide'}, status=400)

            relais_obj.save()

            return redirect('relais-control')  # Redirige vers la même page pour voir l'état mis à jour

        except Relais.DoesNotExist:
            return JsonResponse({'error': 'Relais non trouvé'}, status=404)

    # Si la requête est en GET, afficher les relais existants et leurs états
    return render(request, 'espcontrol/control_relais.html', {'relais': relais})





@login_required
def dashboard(request):
    comptages = Comptage.objects.filter(user=request.user).order_by('-timestamp')[:5]
    return render(request, 'espcontrol/dashboard.html', {'comptages': comptages})


@login_required
def comptage_live(request):
    comptages = Comptage.objects.all().order_by('-timestamp')[:20]
    data = [
        {
            "id": c.id,
            "compteur": c.compteur,
            "timestamp": c.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        }
        for c in comptages
    ]
    return JsonResponse(data, safe=False)



from django.db import connection
from django.contrib import messages
import time

@login_required
def initCompteur(request):
    if request.method == 'POST' and 'reset' in request.POST:

        # 1) Supprimer toutes les données
        Comptage.objects.all().delete()

        # Petite pause pour être sûr que tout soit bien écrit
        time.sleep(0.3)

        # 2) Réinitialiser l'auto-increment
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM sqlite_sequence WHERE name='espcontrol_comptage';")

        # 3) Message de succès
        messages.success(request, "Compteur réinitialisé avec succès !")

        return redirect('dashboard')

    comptages = Comptage.objects.all().order_by('-timestamp')[:5]
    return render(request, 'espcontrol/deleteComp.html', {'comptage': comptages})





import requests
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Relais
from .forms import RelaisForm
from django.contrib.auth.decorators import login_required


from django.db import IntegrityError

@login_required
def control_relais(request):
    if request.method == 'POST':
        form = RelaisForm(request.POST)
        if form.is_valid():
            relais = form.save(commit=False)
            relais.user = request.user
            try:
                relais.save()
                return redirect('afficher_relais')
            except IntegrityError:
                form.add_error('num', 'Ce numéro est déjà utilisé par vous.')
    else:
        form = RelaisForm()

    return render(request, 'espcontrol/control_relais.html', {
        'form': form
    })




@login_required
def afficher_relais(request):
    relais = Relais.objects.filter(user=request.user)
    return render(request, "espcontrol/relais.html", {"relais": relais})


from django.shortcuts import get_object_or_404

@login_required
def toggle_relais(request, num):
    relais = get_object_or_404(Relais, num=num, user=request.user)
    relais.etat = not relais.etat
    relais.save()
    return redirect("afficher_relais")




def display_soil_data(request):
    # Récupérer toutes les données d'humidité
    data = SoilData.objects.all().order_by('-created_at')  # Trier par date décroissante

    # Passer les données au template
    return render(request, 'espcontrol/soil_data.html', {'data': data})



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings  # Pour accéder à DEFAULT_FROM_EMAIL
from .models import SensorData
from datetime import datetime, timedelta
import json

@csrf_exempt
@login_required
def receive_sensor_data(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            temperature = data.get('temperature')
            humidity = data.get('humidity')
            co2 = data.get('co2')
            timestamp_str = data.get('timestamp')

            if not all([temperature, humidity, co2, timestamp_str]):
                return JsonResponse({'status': 'error', 'message': 'Données manquantes'}, status=400)

            # Estimation du type de gaz
            ppm_estimee = co2
            if ppm_estimee > 1000:
                gaz_type = "CO2 ÉLEVÉ"
            elif 25 <= ppm_estimee <= 50:
                gaz_type = "NH3 POSSIBLE"
            elif 0.5 <= ppm_estimee <= 5:
                gaz_type = "BENZÈNE RISQUÉ"
            else:
                gaz_type = "Air sain"

            # Conversion de la date
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")

            # Sauvegarde en base de données
            sensor_data = SensorData(
                temperature=temperature,
                humidity=humidity,
                co2=co2,
                timestamp=timestamp,
                gaz_type=gaz_type,
                user=request.user
            )
            sensor_data.save()

            # Conditions critiques → envoyer un mail
            if temperature > 30 or gaz_type in ["CO2 ÉLEVÉ", "NH3 POSSIBLE", "BENZÈNE RISQUÉ"]:
                # Vérification de l'heure du dernier envoi d'email
                last_email_time = request.user.profile.last_alert_sent if hasattr(request.user, 'profile') else None

                if last_email_time:
                    time_diff = datetime.now() - last_email_time
                    if time_diff < timedelta(minutes=10):  # Limite d'envoi à 10 minutes
                        return JsonResponse({'status': 'success', 'message': 'Alerte déjà envoyée récemment'}, status=200)

                # Envoi de l'email
                send_mail(
                    subject='🚨 Alerte Environnementale',
                    message=(
                        f"Une alerte a été détectée pour l'utilisateur {request.user.username} :\n\n"
                        f"🌡 Température : {temperature} °C\n"
                        f"💧 Humidité    : {humidity} %\n"
                        f"💨 CO2         : {co2} ppm\n"
                        f"⚠️ Type de gaz : {gaz_type}\n"
                        f"🕒 Date        : {timestamp.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                        f"Prenez des mesures immédiates !"
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[request.user.email],
                    fail_silently=False,
                )

                # Mise à jour du dernier envoi d'alerte
                if hasattr(request.user, 'profile'):
                    request.user.profile.last_alert_sent = datetime.now()
                    request.user.profile.save()

            return JsonResponse({'status': 'success', 'message': 'Données enregistrées'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Seules les requêtes POST sont autorisées'}, status=405)


#vue pour la temperature de la thermistance

from .models import NtcSensorData
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.core.mail import send_mail
from django.conf import settings
from django.utils.timezone import now
from datetime import datetime, timedelta
import json

@csrf_exempt
@login_required
def receive_sensor_NTC(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # ✅ si c'est une seule mesure
            if "temperature" in data and "timestamp" in data:
                mesures = [data]
            else:
                # ✅ sinon on attend une liste de mesures
                mesures = data.get("mesures", [])

            if not mesures:
                return JsonResponse({'status': 'error', 'message': 'Aucune donnée reçue'}, status=400)

            saved = 0
            alert_sent = False

            for m in mesures:
                temperature = m.get('temperature')
                timestamp_str = m.get('timestamp')

                if temperature is None or timestamp_str is None:
                    continue

                # Conversion du timestamp
                timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")

                # Sauvegarde en base
                NtcSensorData.objects.create(
                    temperature=temperature,
                    timestamp=timestamp,
                    user=request.user
                )
                saved += 1

                # Vérification d’alerte
                if temperature > 30 and not alert_sent:
                    last_email_time = getattr(request.user.profile, 'last_alert_sent', None)

                    if not last_email_time or (now() - last_email_time) >= timedelta(minutes=10):
                        send_mail(
                            subject='🚨 Alerte Température',
                            message=(
                                f"Température critique détectée pour l'utilisateur {request.user.username} :\n\n"
                                f"🌡 Température : {temperature} °C\n"
                                f"🕒 Date : {timestamp.strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                                f"Prenez des mesures immédiates !"
                            ),
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[request.user.email],
                            fail_silently=False,
                        )

                        if hasattr(request.user, 'profile'):
                            request.user.profile.last_alert_sent = now()
                            request.user.profile.save()

                        alert_sent = True

            return JsonResponse({'status': 'success', 'message': f'{saved} mesures enregistrées'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Seules les requêtes POST sont autorisées'}, status=405)


#Vue pour afficher les donées de la thermistance
from django.contrib.auth.decorators import login_required
from django.db.models import Max, Min
from .models import NtcSensorData

@login_required
def view_ntc_data(request):
    # Toutes les mesures de l'utilisateur, triées par timestamp croissant
    data_list = NtcSensorData.objects.filter(user=request.user).order_by('timestamp')

    # Dernière mesure
    last_temp = data_list.last() if data_list.exists() else None

    # Température max et min
    max_temp = data_list.aggregate(Max('temperature'))['temperature__max'] if data_list.exists() else None
    min_temp = data_list.aggregate(Min('temperature'))['temperature__min'] if data_list.exists() else None

    # Préparer les données pour le graphique (temps en secondes depuis la première mesure)
    labels_sec = []
    temps = []
    if data_list.exists():
        t0 = data_list.first().timestamp
        for d in data_list:
            delta_sec = (d.timestamp - t0).total_seconds()
            labels_sec.append(int(delta_sec))
            temps.append(d.temperature)

    context = {
        'data_list': data_list,
        'last_temp': last_temp,
        'max_temp': max_temp,
        'min_temp': min_temp,
        'labels_sec': labels_sec,  # temps depuis la première mesure
        'temps': temps,            # températures
    }
    return render(request, 'espcontrol/ntc_data.html', context)



#export ecel ntc
# views.py
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from .models import NtcSensorData


@login_required
def export_ntc_excel(request):
    # Récupérer les dates envoyées depuis le formulaire
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Validation des formats de date
    try:
        if start_date:
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        if end_date:
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Format de date invalide. Utilise AAAA-MM-JJ.", status=400)

    # Filtrer les données par utilisateur connecté
    data = NtcSensorData.objects.filter(user=request.user)

    # Si l'utilisateur a choisi une période, on filtre
    if start_date and end_date:
        data = data.filter(timestamp__date__gte=start_date, timestamp__date__lte=end_date)

    # Vérifier s'il y a des données
    if not data.exists():
        return HttpResponse("Aucune donnée trouvée pour cette période.", status=404)

    # Créer le fichier Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Données Thermistance NTC"

    # Ajouter les en-têtes
    headers = ["Date et heure", "Température (°C)"]
    sheet.append(headers)

    # Ajouter les données
    for entry in data.order_by("-timestamp"):
        sheet.append([
            entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            entry.temperature,
        ])

    # Mise en forme (gras + largeur colonnes)
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = 25

    # Créer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ntc_data.xlsx"'

    workbook.save(response)
    return response



#vue pour afficher et calculer les mesures
from .models import SensorData
from django.db.models import Avg, Min, Max

@login_required
def gas_data_view(request):
    # Récupérer les données pour l'utilisateur connecté
    data = SensorData.objects.filter(user=request.user).order_by('-timestamp')

    # Calcul des statistiques
    stats = data.aggregate(
        avg_temp=Avg('temperature'),
        min_temp=Min('temperature'),
        max_temp=Max('temperature'),
        avg_hum=Avg('humidity'),
        min_hum=Min('humidity'),
        max_hum=Max('humidity'),
        avg_co2=Avg('co2'),
        min_co2=Min('co2'),
        max_co2=Max('co2')
    )

    return render(request, 'espcontrol/gas_data.html', {'data': data, 'stats': stats})




import openpyxl
from django.http import HttpResponse

#Vues pour télécharger les données recoltées par le capteur d'humidité DHT11

@login_required
@plan_required(['basic', 'pro'])
def export_excel(request):
    # Récupérer les dates envoyées depuis le formulaire
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Filtrer les données par utilisateur connecté
    data = DHTData.objects.filter(user=request.user)

    # Si l'utilisateur a choisi une période, on filtre
    if start_date and end_date:
        data = data.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    # Créer le fichier Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Données DHT11"

    # Ajouter les en-têtes
    sheet.append(["Date", "Température (°C)", "Humidité (%)"])

    # Ajouter les données
    for entry in data.order_by("-created_at"):
        sheet.append([
            entry.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            entry.temperature,
            entry.humidity
        ])

    # Créer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="dht11_data.xlsx"'

    workbook.save(response)
    return response

# Importe ta fonction de détection d'anomalies ici



def detect_anomalies(device, window=50, sigma_threshold=2):
    # Récupérer les données pour ce device spécifique
    data_points = list(AppareilData.objects.filter(device=device).order_by('-received_at')[:window])
    if not data_points:
        return []

    # Liste des capteurs SDS011 et autres
    sensors = ['pm2p5', 'pm10', 'temperature', 'humidity', 'mq135_ppm']
    anomalies = []

    for sensor in sensors:
        # Extraire uniquement les valeurs numériques
        values = [d.payload.get(sensor) for d in data_points
                 if d.payload and isinstance(d.payload.get(sensor), (int, float))]

        if len(values) < 5: # Besoin d'un minimum de points pour la moyenne
            continue

        mean = np.mean(values)
        std = np.std(values)
        latest_data = data_points[0]
        latest_value = latest_data.payload.get(sensor)

        if latest_value is not None:
            # Algorithme du Z-Score
            if std > 0 and (abs(latest_value - mean) > sigma_threshold * std):
                anomalies.append({
                    'sensor': sensor,
                    'value': latest_value,
                    'mean': round(float(mean), 2),
                    'std': round(float(std), 2)
                })
                # Marquer en base de données
                latest_data.is_anomaly = True
                latest_data.save()

                # --- Création automatique de l'alerte pour l'Agent IA ---
                AgentAlert.objects.get_or_create(
                    user=device.user,
                    device=device,
                    sensor=sensor,
                    value=str(latest_value),
                    level='CRITICAL' if sensor.startswith('pm') and latest_value > 50 else 'WARN',
                    message=f"Anomalie détectée sur {sensor}: {latest_value} (Moyenne: {round(mean, 1)})",
                    is_read=False
                )
    return anomalies

@login_required
def air_quality_dashboard(request):
    """
    Vue principale du Dashboard Founatek Nexus.
    Affiche les alertes IA, les statistiques et les graphiques par appareil.

    🎯 COHÉRENCE AGENT : le badge AQI utilise la même logique que l'agent IA
    (moyenne glissante 5 mesures) — un pic ponctuel (cigarette) ne déclenche
    pas un changement de statut alarmant.
    """
    from espcontrol.models import Relais

    now = timezone.now()
    yesterday = now - timedelta(hours=24)

    # 1. Récupération des alertes non lues récentes (30 dernières minutes)
    recent_limit = now - timedelta(minutes=30)
    agent_alerts = AgentAlert.objects.filter(
        user=request.user,
        is_read=False,
        created_at__gte=recent_limit
    ).order_by("-created_at")

    devices = Device.objects.filter(user=request.user)
    devices_data = []
    sensors_list = ['pm2p5', 'pm10', 'temperature', 'humidity', 'mq135_ppm']

    for device in devices:
        # Récupération des données brutes
        device_readings = AppareilData.objects.filter(device=device).order_by('-received_at')
        all_points      = device_readings[:50]
        recent_readings = device_readings.filter(received_at__gte=yesterday)

        # 2. Détection d'anomalies (IA Z-Score)
        anomalies = detect_anomalies(device, window=5)

        # 3. Calcul des statistiques PM2.5 et PM10 (sur 24h)
        pm25_v = [
            r.payload.get('pm2p5') for r in recent_readings
            if r.payload and isinstance(r.payload.get('pm2p5'), (int, float))
        ]
        pm10_v = [
            r.payload.get('pm10') for r in recent_readings
            if r.payload and isinstance(r.payload.get('pm10'), (int, float))
        ]
        stats = {
            'avg_pm25': round(sum(pm25_v) / len(pm25_v), 2) if pm25_v else 0,
            'avg_pm10': round(sum(pm10_v) / len(pm10_v), 2) if pm10_v else 0,
        }

        # ════════════════════════════════════════════════════════════════
        # 4. Calcul du statut AQI — MOYENNE GLISSANTE 5 MESURES
        # ════════════════════════════════════════════════════════════════
        # Cohérent avec la logique de l'agent IA : on filtre les pics
        # ponctuels (cigarette, voiture) pour ne réagir qu'aux tendances
        # durables de pollution.
        latest = device_readings.first()
        aqi = {"status": "N/A", "color": "#64748b", "icon": "❓"}

        # Moyenne glissante sur les 5 dernières mesures
        recent_5 = list(device_readings[:5])
        pm25_values = [
            r.payload.get('pm2p5') for r in recent_5
            if r.payload and isinstance(r.payload.get('pm2p5'), (int, float))
        ]

        if pm25_values:
            pm25_moy = sum(pm25_values) / len(pm25_values)

            # Classification selon les seuils OMS 2021
            if pm25_moy <= 15:
                aqi = {
                    "status": "Bon",
                    "color":  "#10b981",
                    "icon":   "✅",
                    "value":  round(pm25_moy, 1),
                }
            elif pm25_moy <= 35:
                aqi = {
                    "status": "Modéré",
                    "color":  "#f59e0b",
                    "icon":   "⚠️",
                    "value":  round(pm25_moy, 1),
                }
            elif pm25_moy <= 55:
                aqi = {
                    "status": "Mauvais",
                    "color":  "#ef4444",
                    "icon":   "🔴",
                    "value":  round(pm25_moy, 1),
                }
            else:
                aqi = {
                    "status": "Très mauvais",
                    "color":  "#7c2d12",
                    "icon":   "☠️",
                    "value":  round(pm25_moy, 1),
                }

        # 5. Préparation des points pour les graphiques Chart.js
        processed_points = []
        for p in all_points:
            processed_points.append({
                "id":          p.id,
                "received_at": p.received_at.isoformat(),
                "payload":     p.payload,
                "is_anomaly":  p.is_anomaly,
            })

        # 6. Compilation des données par appareil
        devices_data.append({
            "device":         device,
            "latest_reading": latest,
            "stats":          stats,
            "aqi_status":     aqi,
            "chart_data":     list(reversed(processed_points[:20])),
            "data_points":    processed_points,
            "has_anomaly":    bool(anomalies),
            "total_count":    AppareilData.objects.filter(device=device).count(),
            "relais_list":    Relais.objects.filter(user=request.user),
        })

    return render(request, 'espcontrol/air_quality_dashboard.html', {
        "devices_data":  devices_data,
        "sensors_list":  sensors_list,
        "agent_alerts":  agent_alerts,
    })

@api_permission_required
@api_view(['GET'])
def air_quality_data_api(request):
    """API basée sur les vraies données IoT (AppareilData)"""

    limit = int(request.GET.get('limit', 50))

    readings = (
        AppareilData.objects
        .select_related("device")
        .order_by('-received_at')[:limit]
    )

    data_by_device = {}

    for r in readings:
        payload = r.payload or {}
        device_name = r.device.name if r.device else "Inconnu"

        if device_name not in data_by_device:
            data_by_device[device_name] = {
                "device": device_name,
                "latest": r,
                "readings": []
            }

        data_by_device[device_name]["readings"].append(r)

    response_data = []

    for device_name, data in data_by_device.items():
        pm25_list = [r.payload.get("pm2p5", 0) for r in data["readings"]]
        pm10_list = [r.payload.get("pm10", 0) for r in data["readings"]]

        avg_pm25 = sum(pm25_list) / len(pm25_list) if pm25_list else 0
        avg_pm10 = sum(pm10_list) / len(pm10_list) if pm10_list else 0

        latest_payload = data["latest"].payload or {}

        response_data.append({
            "device": device_name,
            "latest": {
                "pm2p5": latest_payload.get("pm2p5"),
                "pm10": latest_payload.get("pm10"),
                "timestamp": data["latest"].received_at.isoformat(),
            },
            "averages": {
                "pm2p5": round(avg_pm25, 1),
                "pm10": round(avg_pm10, 1),
            },
            "location": {
                "lat": latest_payload.get("latitude"),
                "lon": latest_payload.get("longitude"),
            }
        })

    return Response({"data": response_data}, status=status.HTTP_200_OK)


import openpyxl
from django.http import HttpResponse



@login_required
def export_air_founatek_nexus_excel(request, device_id):

    # ──────────────────────────────────────────────────────────
    # SÉCURITÉ
    # ──────────────────────────────────────────────────────────
    device = get_object_or_404(
        Device,
        id=device_id,
        user=request.user
    )

    # ──────────────────────────────────────────────────────────
    # FILTRAGE PAR PÉRIODE (optionnel)
    # ──────────────────────────────────────────────────────────
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    query = AppareilData.objects.filter(device=device)

    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').replace(hour=0, minute=0)
            end = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59)
            query = query.filter(received_at__range=[start, end])
        except ValueError:
            pass

    # ──────────────────────────────────────────────────────────
    # DONNÉES
    # ──────────────────────────────────────────────────────────
    data = query.order_by('-received_at')[:2000]

    # ──────────────────────────────────────────────────────────
    # 🚨 DÉTECTION ANOMALIES
    # ──────────────────────────────────────────────────────────
    anomalies = detect_anomalies(device, window=100, sigma_threshold=2)
    anomaly_ids = set(a.get('id') for a in anomalies if 'id' in a)

    # ──────────────────────────────────────────────────────────
    # 🤖 ALERTES AGENT IA
    # ──────────────────────────────────────────────────────────
    agent_alerts = AgentAlert.objects.filter(
        user=request.user,
        device=device
    ).order_by('-created_at')[:50]

    alert_map = {a.id: a for a in agent_alerts}

    # ──────────────────────────────────────────────────────────
    # CLASSEUR EXCEL
    # ──────────────────────────────────────────────────────────
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Mesures {device.name}"

    # ──────────────────────────────────────────────────────────
    # EN-TÊTES
    # ──────────────────────────────────────────────────────────
    sheet.append([
        "Date/Heure",
        "Appareil",
        "PM2.5 (µg/m³)",
        "PM10 (µg/m³)",
        "Gaz MQ135 (PPM)",
        "Température (°C)",
        "Humidité (%)",
        "Latitude",
        "Longitude",
        "Satellites GPS",
        "🚨 Anomalie",
        "🤖 Alerte Agent"
    ])

    # Style header
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # ──────────────────────────────────────────────────────────
    # DONNÉES LIGNES
    # ──────────────────────────────────────────────────────────
    for entry in data:
        p = entry.payload or {}

        is_anomaly = entry.is_anomaly or (entry.id in anomaly_ids)
        anomaly_label = "⚠️ OUI" if is_anomaly else "Non"

        alert_msg = ""
        if entry.id in alert_map:
            alert = alert_map[entry.id]
            alert_msg = f"{alert.level}: {alert.message}"

        row = [
            entry.received_at.strftime("%d/%m/%Y %H:%M:%S"),
            device.name,
            p.get("pm2p5", "0"),
            p.get("pm10", "0"),
            p.get("mq135_ppm", "0"),
            p.get("temperature", "0"),
            p.get("humidity", "0"),
            p.get("latitude", "0"),
            p.get("longitude", "0"),
            p.get("satellites", "0"),
            anomaly_label,
            alert_msg
        ]

        sheet.append(row)

        if is_anomaly:
            anomaly_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            for cell in sheet[sheet.max_row]:
                cell.fill = anomaly_fill

    # ──────────────────────────────────────────────────────────
    # STYLE COLONNES
    # ──────────────────────────────────────────────────────────
    column_widths = {
        "A": 22, "B": 25, "C": 15, "D": 15, "E": 18,
        "F": 18, "G": 15, "H": 16, "I": 16, "J": 16,
        "K": 15, "L": 35
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # ──────────────────────────────────────────────────────────
    # FEUILLE STATISTIQUES
    # ──────────────────────────────────────────────────────────
    stats_sheet = workbook.create_sheet("Statistiques")

    pm25_list = [d.payload.get('pm2p5', 0) for d in data if d.payload and d.payload.get('pm2p5')]
    pm10_list = [d.payload.get('pm10', 0) for d in data if d.payload and d.payload.get('pm10')]

    stats_sheet.append(["Métrique", "Valeur"])
    stats_sheet.append(["Total mesures", len(data)])
    stats_sheet.append(["Anomalies détectées", len(anomalies)])
    stats_sheet.append(["Alertes Agent", agent_alerts.count()])
    stats_sheet.append(["PM2.5 Moyenne", round(sum(pm25_list) / len(pm25_list), 2) if pm25_list else 0])
    stats_sheet.append(["PM2.5 Min", min(pm25_list) if pm25_list else 0])
    stats_sheet.append(["PM2.5 Max", max(pm25_list) if pm25_list else 0])
    stats_sheet.append(["PM10 Moyenne", round(sum(pm10_list) / len(pm10_list), 2) if pm10_list else 0])
    stats_sheet.append(["PM10 Min", min(pm10_list) if pm10_list else 0])
    stats_sheet.append(["PM10 Max", max(pm10_list) if pm10_list else 0])

    for cell in stats_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    stats_sheet.column_dimensions["A"].width = 25
    stats_sheet.column_dimensions["B"].width = 20

    # ──────────────────────────────────────────────────────────
    # NOM FICHIER
    # ──────────────────────────────────────────────────────────
    filename = (
        f"Data_{device.name.replace(' ', '_')}_"
        f"{timezone.now().strftime('%d_%m_%Y_%Hh%M')}.xlsx"
    )

    # ──────────────────────────────────────────────────────────
    # RÉPONSE HTTP
    # ──────────────────────────────────────────────────────────
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    workbook.save(response)

    return response

# Vues pour télécharger les données recoltées par le capteur de gaz MQTT35

@login_required
def export_gas_excel(request):

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    data = SensorData.objects.filter(user=request.user)

    if start_date and end_date:
        data = data.filter(
            timestamp__date__gte=start_date,
            timestamp__date__lte=end_date
        )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Données Gaz"

    sheet.append(["Date", "Température (°C)", "Humidité (%)", "CO2 (ppm)", "Type de Gaz"])

    for entry in data.order_by("-timestamp"):
        sheet.append([
            entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            entry.temperature,
            entry.humidity,
            entry.co2,
            entry.gaz_type
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="gas_data.xlsx"'

    workbook.save(response)
    return response

@login_required
def export_air_data_excel(request):
    # Récupérer les données des appareils de l'utilisateur
    # On prend les 1000 dernières mesures par exemple
    devices = Device.objects.filter(user=request.user)
    data = AppareilData.objects.filter(device__in=devices).order_by("-received_at")[:1000]

    # Création du classeur Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Données Air Quality"

    # En-têtes des colonnes
    headers = [
        "Horodatage", "Appareil", "PM2.5 (µg/m³)", "PM10 (µg/m³)",
        "Température (°C)", "Humidité (%)", "Gaz (PPM)", "Latitude", "Longitude"
    ]
    sheet.append(headers)

    # Remplissage des lignes
    for entry in data:
        p = entry.payload or {}
        sheet.append([
            entry.received_at.strftime("%d/%m/%Y %H:%M:%S"),
            entry.device.name,
            p.get("pm2p5", "N/A"),
            p.get("pm10", "N/A"),
            p.get("temperature", "N/A"),
            p.get("humidity", "N/A"),
            p.get("mq135_ppm", "N/A"),
            p.get("latitude", "N/A"),
            p.get("longitude", "N/A"),
        ])

    # Configuration de la réponse HTTP pour le téléchargement
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Founatek_Air_Data_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response



# views.py pour la bande de led ws2812b
from django.http import JsonResponse
from .models import LEDColor

@login_required
def led_color_esp(request):
    user = request.user
    color = LEDColor.objects.filter(user=user).last()
    if color:
        data = {"r": color.r, "g": color.g, "b": color.b}
    else:
        data = {"r": 0, "g": 0, "b": 0}
    return JsonResponse(data)


# views.py
import json
import logging
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .chatbot_model import Chatbot

logger = logging.getLogger(__name__)

@login_required
def chatbot_view(request):
    if request.method != "POST":
        return JsonResponse({"reponse": "Méthode non autorisée."}, status=405)

    try:
        data    = json.loads(request.body or "{}")
        raw_msg = str(data.get("message", "")).strip()
    except json.JSONDecodeError:
        return JsonResponse({"reponse": "⚠️ Données invalides."})

    if not raw_msg:
        return JsonResponse({"reponse": "Dis-moi ce que tu veux faire 🙂"})

    bot = Chatbot(request.user)

    try:
        response = bot.get_response(raw_msg)
    except Exception as e:
        logger.error(f"Erreur chatbot user={request.user.id}: {e}")
        return JsonResponse({"reponse": f"⚠️ Erreur interne : {e}"})

    # ── CORRECTION BUG "aide" ────────────────────────────────
    # Si le bot retourne un dict avec boutons, s'assurer
    # que "reponse" est toujours présent
    if isinstance(response, dict):
        if "reponse" not in response:
            response["reponse"] = "Voici les options :"
        return JsonResponse(response)

    return JsonResponse({"reponse": response, "tts": response})


#Vue pour le contrôle d'accès automatique
# Vue pour le contrôle d'accès automatique
from django.http import JsonResponse
from django.utils import timezone
from .models import Badge, Door, AccessRule, AccessLog
from .utils import api_permission_required
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@api_permission_required
def check_access(request):
    if request.method != "POST":
        return JsonResponse({"error": "Méthode non autorisée"}, status=405)

    uid = request.POST.get("uid")
    door_slug = request.POST.get("door")
    ip = request.META.get("REMOTE_ADDR")

    if not uid or not door_slug:
        return JsonResponse({"allowed": False, "message": "UID et porte requis"}, status=400)

    # Vérifie que la porte existe (supprime le filtre user=request.user)
    try:
        door = Door.objects.get(slug=door_slug, is_active=True)
    except Door.DoesNotExist:
        return JsonResponse({"allowed": False, "message": "Porte inconnue"}, status=404)

    # Vérifie le badge actif
    try:
        badge = Badge.objects.get(uid=uid, is_active=True)
        owner_name = badge.owner.username if badge.owner else "inconnu"
    except Badge.DoesNotExist:
        AccessLog.objects.create(
            uid=uid,
            door=door,
            user=None,
            allowed=False,
            ip_address=ip,
            raw_payload=request.POST.dict()
        )
        return JsonResponse({
            "allowed": False,
            "message": "Badge inconnu",
            "action": "deny",
            "owner": "inconnu",
            "timestamp": timezone.now()
        }, status=403)

    # Vérifie les règles d'accès
    rules = AccessRule.objects.filter(badge=badge, door=door)
    access_allowed = any(rule.is_currently_valid() for rule in rules)

    # Log
    AccessLog.objects.create(
        badge=badge,
        uid=uid,
        door=door,
        user=badge.owner,
        allowed=access_allowed,
        ip_address=ip,
        raw_payload=request.POST.dict()
    )

    response = {
        "allowed": access_allowed,
        "message": f"Accès {'autorisé' if access_allowed else 'refusé'} pour {owner_name}",
        "action": "open" if access_allowed else "deny",
        "owner": owner_name,
        "timestamp": timezone.now()
    }

    return JsonResponse(response, status=200 if access_allowed else 403)


from django.core.paginator import Paginator
from django.db.models import Count, Case, When, IntegerField, Q
from django.db.models import Count, Case, When, IntegerField, Q
from .models import AccessLog

@login_required
def access_logs_view(request):
    # On filtre les logs pour ne garder que ceux de l'utilisateur connecté
    logs = AccessLog.objects.select_related('badge', 'door', 'user') \
                            .filter(user=request.user) \
                            .order_by('-timestamp')

    # Récupérer filtres GET facultatifs
    door_filter = request.GET.get("door")
    badge_filter = request.GET.get("badge")

    # Appliquer filtres
    if door_filter:
        logs = logs.filter(door__name__icontains=door_filter)
    if badge_filter:
        logs = logs.filter(Q(badge__uid__icontains=badge_filter) |
                           Q(badge__label__icontains=badge_filter))

    # Pagination
    page_number = request.GET.get("page", 1)
    paginator = Paginator(logs, 20)  # 20 logs par page
    page_obj = paginator.get_page(page_number)

    # Comptages globaux
    total_logs = logs.count()
    total_allowed = logs.filter(allowed=True).count()
    total_denied = logs.filter(allowed=False).count()

    context = {
        "logs": page_obj,  # page_obj contient seulement les logs de l'utilisateur
        "page_obj": page_obj,
        "total_logs": total_logs,
        "total_allowed": total_allowed,
        "total_denied": total_denied,
        "door_filter": door_filter or "",
        "badge_filter": badge_filter or "",
    }

    return render(request, "espcontrol/access_logs.html", context)


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.authtoken.models import Token
from .serializers import AppareilDataSerializer
import numpy as np
# ...existing code...
from django.utils import timezone
from espcontrol.models import ActionLog

# Limite maximale du body d'ingestion (bytes) pour protéger contre les payloads trop volumineux
MAX_INGEST_BODY_BYTES = 16 * 1024  # 16 KB


class SensorIngestAPIView(APIView):
    """
    API d'ingestion des données ESP32.

    🎯 NOUVEAU : Déclenchement événementiel de l'agent IA
    L'agent FOUNATEK analyse la mesure IMMÉDIATEMENT après réception,
    dans la même requête HTTP. Plus besoin de polling 0.5s.

    Délai total ESP32 → Django → Agent → Alerte ≈ 100ms
    """
    authentication_classes = []
    permission_classes = []

    def post(self, request):
        # Protection: refuser les payloads trop volumineux
        try:
            content_length = int(request.META.get('CONTENT_LENGTH') or 0)
        except Exception:
            content_length = 0
        if content_length and content_length > MAX_INGEST_BODY_BYTES:
            return Response({'error': 'Payload trop volumineux'}, status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE)
        # fallback: check raw body length
        try:
            if not content_length and request.body and len(request.body) > MAX_INGEST_BODY_BYTES:
                return Response({'error': 'Payload trop volumineux'}, status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE)
        except Exception:
            pass

        # Déterminer l'utilisateur (session ou token)
        user = None
        if hasattr(request, "user") and getattr(request.user, "is_authenticated", False):
            user = request.user

        token_key = None
        auth_header = request.META.get("HTTP_AUTHORIZATION", "") or ""
        if isinstance(auth_header, str) and auth_header.startswith("Token "):
            token_key = auth_header.split(" ", 1)[1]
            try:
                token = Token.objects.get(key=token_key)
                if not user:
                    user = token.user
            except Token.DoesNotExist:
                pass

        # Basic structural validation before serializer
        if not isinstance(request.data, dict):
            return Response({'error': 'Payload JSON invalide'}, status=status.HTTP_400_BAD_REQUEST)
        if 'device_id' not in request.data or 'data' not in request.data:
            return Response({'error': 'Champs requis manquants (device_id, data)'}, status=status.HTTP_400_BAD_REQUEST)
        # ensure payload is an object/dict
        if not isinstance(request.data.get('data'), (dict,)):
            return Response({'error': 'Le champ data doit être un objet JSON'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = AppareilDataSerializer(
            data=request.data,
            context={'user': user, 'token_key': token_key}
        )
        if serializer.is_valid():
            appdata = serializer.save()

            device = getattr(appdata, "device", None)
            if device:
                device.last_seen = timezone.now()
                device.save(update_fields=["last_seen"])

            # ActionLog conforme au modèle (user + action + details JSON)
            if user:
                ActionLog.objects.create(
                    user=user,
                    action="ingest",
                    details={
                        "device_id": getattr(device, "device_id", None),
                        "payload": getattr(appdata, "payload", None),
                        "token_key": token_key,
                    },
                )

            # ════════════════════════════════════════════════════════════
            # 🎯 DÉCLENCHEMENT ÉVÉNEMENTIEL DE L'AGENT IA
            # ════════════════════════════════════════════════════════════
            # L'agent analyse la nouvelle mesure IMMÉDIATEMENT.
            # Si une pollution est détectée, l'alerte est créée AVANT
            # même que l'ESP32 reçoive sa réponse 201.
            # ════════════════════════════════════════════════════════════
            if user:
                try:
                    from espcontrol.agent.agent import FounatekAgent
                    agent = FounatekAgent(user)
                    agent.run()
                except Exception as e:
                    # L'agent ne doit JAMAIS bloquer l'ingestion ESP32
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(
                        f"⚠️ Agent IA failed silently for user={user.username}: {e}"
                    )

            return Response({"status": "data received"}, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@login_required
def dashboard_univers(request):
    # On récupère tous les appareils de l'utilisateur
    devices = Device.objects.filter(user=request.user)
    devices_data = []

    # Liste exhaustive des capteurs possibles
    sensors_list = ['temperature', 'humidity', 'soil_moisture', 'ultrason', 'mq135_ppm']

    for device in devices:
        # On récupère les 50 derniers points
        data_points_qs = AppareilData.objects.filter(device=device).order_by('-received_at')[:50]

        # Détection des anomalies (ta fonction existante)
        anomalies = detect_anomalies(device, window=50, sigma_threshold=2)
        anomaly_ids = [a.get('id') for a in anomalies if 'id' in a]

        # On prépare les données pour le template
        processed_points = []
        for p in data_points_qs:
            point_dict = {
                "id": p.id,
                "received_at": p.received_at.isoformat(), # Pour faciliter le JS
                "payload": p.payload,
                "is_anomaly": p.id in anomaly_ids
            }
            processed_points.append(point_dict)

        devices_data.append({
            "device": device,
            "data_points": list(reversed(processed_points)), # Chronologique pour Chart.js
            "has_anomaly": bool(anomalies),
            "anomalies": anomalies,
        })

    alerts = AgentAlert.objects.filter(
        user=request.user,
        is_read=False
    ).order_by('-created_at')[:20]

    return render(request, "espcontrol/dashboardUniv.html", {
        "devices_data": devices_data,
        "sensors_list": sensors_list,
        "agent_alerts": alerts
    })






def detect_anomalies(device, window=50, sigma_threshold=2):
    data_points = list(AppareilData.objects.filter(device=device).order_by('-received_at')[:window])
    if not data_points:
        return []

    sensors = ['temperature', 'humidity', 'soil_moisture', 'ultrason', 'mq135_ppm']
    anomalies = []

    for sensor in sensors:
        values = [d.payload.get(sensor) for d in data_points if sensor in d.payload]
        if len(values) < 5:
            continue

        mean = np.mean(values)
        std = np.std(values)
        latest_value = values[0]
        latest_data = data_points[0]

        if latest_value < mean - sigma_threshold * std or latest_value > mean + sigma_threshold * std:
            anomalies.append({
                'sensor': sensor,
                'value': latest_value,
                'mean': mean,
                'std': std
            })
            # Marquer l'objet le plus récent comme anomalie
            latest_data.is_anomaly = True
            latest_data.save()

    return anomalies



# --- API JSON pour le JS ---
from rest_framework.authentication import SessionAuthentication, TokenAuthentication

class DeviceLast10APIView(APIView):
    authentication_classes = [SessionAuthentication, TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, device_id):
        try:
            device = Device.objects.get(id=device_id, user=request.user)
        except Device.DoesNotExist:
            return Response({"error": "Device introuvable"}, status=status.HTTP_404_NOT_FOUND)

        data_points = list(
            AppareilData.objects.filter(device=device)
            .order_by('-received_at')[:200]
        )
        data_points = list(reversed(data_points))

        result = []
        for d in data_points:
            result.append({
                "time":        d.received_at.strftime('%H:%M:%S'),
                "pm25":        d.payload.get('pm2p5'),
                "pm10":        d.payload.get('pm10'),
                "gas":         d.payload.get('mq135_ppm'),
                "temperature": d.payload.get('temperature'),
                "humidity":    d.payload.get('humidity'),
                "id":          d.id,
            })

        return Response({"readings": result})
#Api pour charger l'agent api depuis le javascript


@login_required
def latest_alerts(request):
    # 1️⃣ récupérer les IDs des alertes non lues (limitées)
    alert_ids = list(
        AgentAlert.objects
        .filter(user=request.user, is_read=False)
        .order_by('-created_at')
        .values_list("id", flat=True)[:10]
    )

    # 2️⃣ récupérer les objets complets
    alerts = AgentAlert.objects.filter(id__in=alert_ids).order_by('-created_at')

    data = []
    for alert in alerts:
        data.append({
            "id": alert.id,
            "message": alert.message,
            "level": alert.level,
            "code": alert.code,
            "created_at": alert.created_at.isoformat(),
            "device": {
                "id": alert.device.id,
                "name": alert.device.name,
                "device_id": alert.device.device_id
            } if alert.device else None
        })

    # 3️⃣ MAINTENANT on peut updater
    AgentAlert.objects.filter(id__in=alert_ids).update(is_read=True)

    return JsonResponse(data, safe=False)


#api pour l'historique des alerts
@login_required
def alerts_by_sensor(request, sensor):
    alerts = (
        AgentAlert.objects
        .filter(user=request.user, sensor=sensor)
        .order_by('-created_at')[:100]
    )

    data = [{
        "message": a.message,
        "value": a.value,
        "level": a.level,
        "created_at": a.created_at.isoformat(),
        "device": a.device.device_id if a.device else None
    } for a in alerts]

    return JsonResponse(data, safe=False)


#api pour les graphes alerts
@login_required
def alert_graph_data(request, sensor):
    alerts = (
        AgentAlert.objects
        .filter(user=request.user, sensor=sensor)
        .order_by('created_at')[:200]
    )

    return JsonResponse({
        "labels": [a.created_at.strftime("%H:%M") for a in alerts],
        "values": [a.value for a in alerts],
        "levels": [a.level for a in alerts],
    })

def menu_page(request):
    return render(request, "espcontrol/menu_page.html")




@csrf_exempt
def mark_alert_read(request, alert_id):
    if request.method == 'POST':
        try:
            alert = AgentAlert.objects.get(id=alert_id, user=request.user)
            alert.is_read = True
            alert.save()
            return JsonResponse({'status': 'success'})
        except AgentAlert.DoesNotExist:
            return JsonResponse({'status': 'error'}, status=404)