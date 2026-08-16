import json
from datetime import datetime, timedelta
from io import BytesIO

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.mail import send_mail
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from PIL import Image
from rest_framework import status
from rest_framework.authentication import SessionAuthentication, TokenAuthentication
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from django.db.models import Avg, Max, Min
from .models import (
    AccessLog,
    AccessRule,
    ActionLog,
    AgentAlert,
    AppareilData,
    Badge,
    Comptage,
    Door,
    DHTData,
    Device,
    LEDColor,
    NtcSensorData,
    Relais,
    SensorData,
    SensorRule,
    SoilData,
    UploadedImage,
)
from .serializers import RelaisSerializer
from .utils import api_permission_required


from rest_framework.authtoken.models import Token

def _json_error(message, status_code=400, **extra):
    payload = {"error": message}
    payload.update(extra)
    return JsonResponse(payload, status=status_code)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_relay_list(request):
    relays = Relais.objects.filter(user=request.user).order_by("num")
    payload = [
        {
            "id": relay.id,
            "num": relay.num,
            "nom": relay.nom,
            "etat": "on" if relay.etat else "off",
        }
        for relay in relays
    ]
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def mobile_relay_detail(request, num):
    try:
        relay = Relais.objects.get(num=num, user=request.user)
    except Relais.DoesNotExist:
        return _json_error("Relais introuvable", 404)

    if request.method == "GET":
        return Response(
            {
                "id": relay.id,
                "num": relay.num,
                "nom": relay.nom,
                "etat": "on" if relay.etat else "off",
            },
            status=status.HTTP_200_OK,
        )

    if request.method == "POST":
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return _json_error("Données JSON invalides", 400)

        state = payload.get("state")
        if state in {"on", "off"}:
            relay.etat = state == "on"
            relay.save(update_fields=["etat"])
            return Response(
                {"num": relay.num, "etat": "on" if relay.etat else "off"},
                status=status.HTTP_200_OK,
            )

        if payload.get("toggle") is True:
            relay.etat = not relay.etat
            relay.save(update_fields=["etat"])
            return Response(
                {"num": relay.num, "etat": "on" if relay.etat else "off"},
                status=status.HTTP_200_OK,
            )

        return _json_error("État non fourni", 400)

    return _json_error("Méthode non autorisée", 405)


@csrf_exempt
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def mobile_soil_data(request):
    if request.method == "POST":
        humidity = request.data.get("humidity")
        if humidity is None:
            return _json_error("Missing humidity data", 400)
        try:
            humidity = int(humidity)
        except (TypeError, ValueError):
            return _json_error("Invalid humidity value", 400)

        SoilData.objects.create(humidity=humidity, user=request.user)
        if humidity < 30:
            send_mail(
                "Alerte : Humidité du sol faible",
                f"L'humidité actuelle du sol est de {humidity}%, ce qui est en dessous du seuil critique de 30%. Pensez à arroser vos plantes.",
                settings.EMAIL_HOST_USER,
                ["isaacdiallo30@gmail.com"],
                fail_silently=False,
            )
        return Response({"status": "success", "humidity": humidity}, status=status.HTTP_201_CREATED)

    records = SoilData.objects.filter(user=request.user).order_by("-created_at")[:20]
    payload = [{"humidity": item.humidity, "created_at": item.created_at.isoformat()} for item in records]
    return Response({"status": "success", "data": payload}, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_sensor_ingest(request):
    payload = request.data if isinstance(request.data, dict) else {}
    sensor_type = str(payload.get("sensor_type", "gas")).lower()

    if sensor_type in {"gas", "dht", "environment"}:
        temperature = payload.get("temperature")
        humidity = payload.get("humidity")
        co2 = payload.get("co2")
        timestamp_str = payload.get("timestamp")
        if not all([temperature is not None, humidity is not None, co2 is not None, timestamp_str]):
            return _json_error("Données manquantes", 400)

        if isinstance(co2, str):
            try:
                co2 = int(co2)
            except ValueError:
                co2 = 0
        if isinstance(temperature, str):
            try:
                temperature = float(temperature)
            except ValueError:
                temperature = 0.0
        if isinstance(humidity, str):
            try:
                humidity = float(humidity)
            except ValueError:
                humidity = 0.0

        ppm_estimee = co2
        if ppm_estimee > 1000:
            gaz_type = "CO2 ÉLEVÉ"
        elif 25 <= ppm_estimee <= 50:
            gaz_type = "NH3 POSSIBLE"
        elif 0.5 <= ppm_estimee <= 5:
            gaz_type = "BENZÈNE RISQUÉ"
        else:
            gaz_type = "Air sain"

        try:
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return _json_error("Format de timestamp invalide", 400)

        SensorData.objects.create(
            temperature=temperature,
            humidity=humidity,
            co2=co2,
            timestamp=timestamp,
            gaz_type=gaz_type,
            user=request.user,
        )
        return Response({"status": "success", "message": "Données enregistrées"}, status=status.HTTP_201_CREATED)

    if sensor_type in {"ntc", "thermistor"}:
        temperature = payload.get("temperature")
        timestamp_str = payload.get("timestamp")
        if temperature is None or not timestamp_str:
            return _json_error("Données manquantes", 400)
        try:
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return _json_error("Format de timestamp invalide", 400)
        NtcSensorData.objects.create(temperature=temperature, timestamp=timestamp, user=request.user)
        return Response({"status": "success", "message": "Mesure NTC enregistrée"}, status=status.HTTP_201_CREATED)

    return _json_error("Type de capteur non pris en charge", 400)


class MobileImageUploadAPIView(APIView):
    authentication_classes = [SessionAuthentication, TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        image_file = request.FILES.get("image") or request.data.get("image")
        if not image_file:
            return Response({"error": "No image provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if hasattr(image_file, "read"):
                image_bytes = image_file.read()
                image_name = getattr(image_file, "name", "received_image.jpg")
            else:
                image_bytes = image_file
                image_name = "received_image.jpg"

            img = Image.open(BytesIO(image_bytes))
            img = img.convert("RGB")
            img_io = BytesIO()
            img.save(img_io, format="JPEG")
            img_io.seek(0)

            uploaded_image = UploadedImage.objects.create(
                image=ContentFile(img_io.read(), image_name),
                user=request.user,
            )
            return Response({"message": "Image enregistrée", "image_id": uploaded_image.id}, status=status.HTTP_201_CREATED)
        except Exception as exc:
            return Response({"error": str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_air_quality_summary(request):
    devices = Device.objects.filter(user=request.user).order_by("name")
    payload = []
    for device in devices:
        latest = AppareilData.objects.filter(device=device).order_by("-received_at").first()
        payload.append(
            {
                "device": {"id": device.id, "name": device.name, "device_id": device.device_id},
                "latest": {
                    "pm2p5":       (latest.payload or {}).get("pm2p5")       if latest else None,
                    "pm10":        (latest.payload or {}).get("pm10")        if latest else None,
                    "mq135_ppm":   (latest.payload or {}).get("mq135_ppm")   if latest else None,
                    "temperature": (latest.payload or {}).get("temperature") if latest else None,
                    "humidity":    (latest.payload or {}).get("humidity")    if latest else None,
                    "latitude":    (latest.payload or {}).get("latitude")    if latest else None,
                    "longitude":   (latest.payload or {}).get("longitude")   if latest else None,
                    "satellites":  (latest.payload or {}).get("satellites")  if latest else None,
                    "timestamp":   latest.received_at.isoformat()            if latest else None,
                },
                "count": AppareilData.objects.filter(device=device).count(),
            }
        )
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_alerts(request):
    show_archived = request.query_params.get("archived", "false") == "true"
    alerts = AgentAlert.objects.filter(user=request.user, is_read=show_archived).order_by("-created_at")[:50]
    payload = [
        {
            "id": alert.id,
            "message": alert.message,
            "level": alert.level,
            "code": alert.code,
            "sensor": alert.sensor,
            "value": alert.value,
            "created_at": alert.created_at.isoformat(),
            "is_read": alert.is_read,
            "device": {"id": alert.device.id, "name": alert.device.name} if alert.device else None,
        }
        for alert in alerts
    ]
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_alerts_archive(request):
    alert_id = request.data.get("id")
    archive_all = request.data.get("all", False)
    if archive_all:
        AgentAlert.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({"archived": "all"})
    if alert_id:
        AgentAlert.objects.filter(id=alert_id, user=request.user).update(is_read=True)
        return Response({"archived": alert_id})
    return Response({"error": "id ou all requis"}, status=400)


@csrf_exempt
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def mobile_alerts_delete(request):
    alert_id = request.data.get("id")
    delete_all = request.data.get("all", False)
    if delete_all:
        AgentAlert.objects.filter(user=request.user, is_read=True).delete()
        return Response({"deleted": "all_archived"})
    if alert_id:
        AgentAlert.objects.filter(id=alert_id, user=request.user).delete()
        return Response({"deleted": alert_id})
    return Response({"error": "id ou all requis"}, status=400)


@csrf_exempt
@api_view(["POST"])
def mobile_access_check(request):
    if request.method != "POST":
        return _json_error("Méthode non autorisée", 405)

    uid = request.POST.get("uid") or request.data.get("uid")
    door_slug = request.POST.get("door") or request.data.get("door")
    ip = request.META.get("REMOTE_ADDR")

    if not uid or not door_slug:
        return _json_error("UID et porte requis", 400)

    try:
        door = Door.objects.get(slug=door_slug, is_active=True)
    except Door.DoesNotExist:
        return Response({"allowed": False, "message": "Porte inconnue"}, status=status.HTTP_404_NOT_FOUND)

    try:
        badge = Badge.objects.get(uid=uid, is_active=True)
        owner_name = badge.owner.username if badge.owner else "inconnu"
    except Badge.DoesNotExist:
        AccessLog.objects.create(uid=uid, door=door, user=None, allowed=False, ip_address=ip, raw_payload=request.POST.dict())
        return Response({"allowed": False, "message": "Badge inconnu", "action": "deny", "owner": "inconnu"}, status=status.HTTP_403_FORBIDDEN)

    rules = AccessRule.objects.filter(badge=badge, door=door)
    allowed = any(rule.is_currently_valid() for rule in rules)

    AccessLog.objects.create(badge=badge, uid=uid, door=door, user=badge.owner, allowed=allowed, ip_address=ip, raw_payload=request.POST.dict())
    return Response(
        {
            "allowed": allowed,
            "message": f"Accès {'autorisé' if allowed else 'refusé'} pour {owner_name}",
            "action": "open" if allowed else "deny",
            "owner": owner_name,
            "timestamp": timezone.now().isoformat(),
        },
        status=status.HTTP_200_OK if allowed else status.HTTP_403_FORBIDDEN,
    )


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_dht_data(request):
    records = DHTData.objects.filter(user=request.user).order_by("-created_at")[:20]
    payload = [
        {
            "id": r.id,
            "temperature": r.temperature,
            "humidity": r.humidity,
            "created_at": r.created_at.isoformat(),
        }
        for r in records
    ]
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_gas_data(request):
    records = SensorData.objects.filter(user=request.user).order_by("-timestamp")[:20]
    payload = [
        {
            "id": r.id,
            "temperature": r.temperature,
            "humidity": r.humidity,
            "co2": r.co2,
            "gaz_type": r.gaz_type,
            "timestamp": r.timestamp.isoformat(),
        }
        for r in records
    ]
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_ntc_data(request):
    records = NtcSensorData.objects.filter(user=request.user).order_by("-timestamp")[:20]
    payload = [
        {
            "id": r.id,
            "temperature": r.temperature,
            "timestamp": r.timestamp.isoformat(),
        }
        for r in records
    ]
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_surveillance(request):
    images = UploadedImage.objects.filter(user=request.user).order_by("-uploaded_at")[:20]
    payload = [
        {
            "id": img.id,
            "url": request.build_absolute_uri(img.image.url) if img.image else None,
            "camera_id": getattr(img, "camera_id", "cam1"),
            "motion_detected": getattr(img, "motion_detected", False),
            "uploaded_at": img.uploaded_at.isoformat(),
        }
        for img in images
    ]
    return Response(payload, status=status.HTTP_200_OK)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_access_logs(request):
    logs = AccessLog.objects.filter(user=request.user).order_by("-timestamp")[:30]
    payload = [
        {
            "id": log.id,
            "uid": log.uid,
            "door": log.door.name if log.door else "—",
            "allowed": log.allowed,
            "timestamp": log.timestamp.isoformat(),
        }
        for log in logs
    ]
    return Response(payload, status=status.HTTP_200_OK)


# ── Mobile Chatbot (Token auth) ───────────────────────────────────────────────
@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_chatbot(request):
    import json as _json
    from .chatbot_model import Chatbot  # noqa: PLC0415
    try:
        data = request.data if isinstance(request.data, dict) else _json.loads(request.body or "{}")
        raw_msg = str(data.get("message", "")).strip()
    except Exception:
        return Response({"reponse": "⚠️ Données invalides."})

    if not raw_msg:
        return Response({"reponse": "Dis-moi ce que tu veux faire 🙂"})

    bot = Chatbot(request.user)
    try:
        response = bot.get_response(raw_msg)
    except Exception as e:
        return Response({"reponse": f"⚠️ Erreur : {e}"})

    if isinstance(response, dict):
        if "reponse" not in response:
            response["reponse"] = "Voici les options :"
        return Response(response)
    return Response({"reponse": response, "tts": response, "buttons": []})


# ── Devices list ──────────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_devices(request):
    devices = Device.objects.filter(user=request.user).order_by("-last_seen")
    payload = []
    for d in devices:
        latest = AppareilData.objects.filter(device=d).order_by("-received_at").first()
        payload.append({
            "id": d.id,
            "name": d.name,
            "device_id": d.device_id,
            "type": getattr(d, "device_type", "unknown"),
            "last_seen": d.last_seen.isoformat() if d.last_seen else None,
            "is_online": (
                (timezone.now() - d.last_seen).total_seconds() < 300
                if d.last_seen else False
            ),
            "reading_count": AppareilData.objects.filter(device=d).count(),
            "latest": latest.payload if latest else {},
        })
    return Response(payload, status=status.HTTP_200_OK)


# ── Stats 24h ─────────────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_stats(request):
    since = timezone.now() - timedelta(hours=24)
    dht  = DHTData.objects.filter(user=request.user, created_at__gte=since)
    gas  = SensorData.objects.filter(user=request.user, timestamp__gte=since)
    ntc  = NtcSensorData.objects.filter(user=request.user, timestamp__gte=since)
    soil = SoilData.objects.filter(user=request.user, created_at__gte=since)

    def agg(qs, field):
        r = qs.aggregate(avg=Avg(field), min=Min(field), max=Max(field))
        return {k: round(v, 2) if v is not None else None for k, v in r.items()}

    return Response({
        "temperature": {**agg(dht,  "temperature"), "count": dht.count()},
        "humidity":    {**agg(dht,  "humidity"),    "count": dht.count()},
        "co2":         {**agg(gas,  "co2"),          "count": gas.count()},
        "ntc_temp":    {**agg(ntc,  "temperature"),  "count": ntc.count()},
        "soil":        {**agg(soil, "humidity"),     "count": soil.count()},
    }, status=status.HTTP_200_OK)


# ── Sensor Rules ──────────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_rules(request):
    rules = SensorRule.objects.filter(user=request.user, active=True)
    payload = [
        {
            "id": r.id,
            "sensor": r.sensor,
            "code": r.code,
            "min_value": r.min_value,
            "max_value": r.max_value,
            "action_type": r.action_type,
            "target_num": getattr(r, "target_num", None),
            "level": r.level,
        }
        for r in rules
    ]
    return Response(payload, status=status.HTTP_200_OK)


# ── Action Logs ───────────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_action_logs(request):
    logs = ActionLog.objects.filter(user=request.user).order_by("-created_at")[:30]
    payload = [
        {
            "id": log.id,
            "action": log.action,
            "details": log.details,
            "created_at": log.created_at.isoformat(),
        }
        for log in logs
    ]
    return Response(payload, status=status.HTTP_200_OK)


# ── LED RGB ───────────────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def mobile_led_color(request):
    if request.method == "POST":
        r_val = int(request.data.get("r", 0))
        g_val = int(request.data.get("g", 0))
        b_val = int(request.data.get("b", 0))
        LEDColor.objects.update_or_create(
            user=request.user,
            defaults={"r": r_val, "g": g_val, "b": b_val},
        )
        return Response({"r": r_val, "g": g_val, "b": b_val})
    led = LEDColor.objects.filter(user=request.user).first()
    return Response({"r": led.r if led else 0, "g": led.g if led else 0, "b": led.b if led else 0})


# ── Comptage ──────────────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_comptage(request):
    latest = Comptage.objects.filter(user=request.user).order_by("-timestamp").first()
    return Response({
        "compteur": latest.compteur if latest else 0,
        "timestamp": latest.timestamp.isoformat() if latest else None,
    })


# ── Création / suppression relais ─────────────────────────────────────────────
@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_relay_create(request):
    num = request.data.get("num")
    nom = request.data.get("nom", f"Relais {num}")
    if num is None:
        return _json_error("Numéro requis", 400)
    try:
        num = int(num)
    except (TypeError, ValueError):
        return _json_error("Numéro invalide", 400)
    if Relais.objects.filter(num=num, user=request.user).exists():
        return _json_error("Ce numéro est déjà utilisé", 409)
    relay = Relais.objects.create(num=num, nom=nom, etat=False, user=request.user)
    return Response({"id": relay.id, "num": relay.num, "nom": relay.nom, "etat": "off"}, status=201)


# ══════════════════════════════════════════════════════════════════════════════
# MODULE FORMATEUR (iot) — gestion des cours
# ══════════════════════════════════════════════════════════════════════════════

def _get_formateur(user):
    from iot.models import FormateurProfile
    return FormateurProfile.objects.filter(user=user).first()


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_formateur_dashboard(request):
    try:
        from iot.models import Parcours, Lecon, Progression, Certification, FormateurProfile
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur introuvable"}, status=403)
        org = fp.organisation
        if org:
            parcours_qs = Parcours.objects.filter(organisation=org)
        else:
            # Formateur sans organisation : montrer tous les parcours publiés
            parcours_qs = Parcours.objects.all()
        total_lecons = Lecon.objects.filter(parcours__in=parcours_qs).count()
        total_apprenants = Progression.objects.filter(lecon__parcours__in=parcours_qs).values("user").distinct().count()
        total_certs = Certification.objects.filter(parcours__in=parcours_qs).count()
        return Response({
            "organisation": org.nom if org else "Formateur",
            "total_parcours": parcours_qs.count(),
            "total_lecons": total_lecons,
            "total_apprenants": total_apprenants,
            "total_certifications": total_certs,
            "parcours": [
                {
                    "id": p.id,
                    "titre": p.titre,
                    "niveau": p.niveau,
                    "is_published": p.is_published,
                    "certifiant": p.certifiant if hasattr(p, 'certifiant') else False,
                    "description": p.description if hasattr(p, 'description') else "",
                    "lecons": [
                        {
                            "id": l.id,
                            "titre": l.titre,
                            "ordre": l.ordre,
                            "resume": l.resume if hasattr(l, 'resume') else "",
                            "quiz_count": l.quizzes.count(),
                            "blocs": list(l.blocs.values("id")),
                        }
                        for l in p.lecons.order_by("ordre").all()
                    ],
                }
                for p in parcours_qs
            ],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_formateur_create_parcours(request):
    try:
        from iot.models import Parcours
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        titre = request.data.get("titre", "")
        description = request.data.get("description", "")
        niveau = request.data.get("niveau", "Débutant")
        certifiant = request.data.get("certifiant", False)
        if not titre:
            return Response({"error": "Titre requis"}, status=400)
        p = Parcours.objects.create(
            organisation=fp.organisation,
            created_by=request.user,
            titre=titre, description=description,
            niveau=niveau, certifiant=certifiant,
            is_published=False,
        )
        return Response({"id": p.id, "titre": p.titre, "slug": p.slug}, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_formateur_edit_parcours(request, parcours_id):
    try:
        from iot.models import Parcours
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        p = Parcours.objects.get(id=parcours_id, organisation=fp.organisation)
        if request.method == "DELETE":
            p.delete()
            return Response({"deleted": True})
        for field in ["titre", "description", "niveau", "certifiant", "is_published"]:
            if field in request.data:
                setattr(p, field, request.data[field])
        p.save()
        return Response({"id": p.id, "titre": p.titre, "is_published": p.is_published})
    except Parcours.DoesNotExist:
        return Response({"error": "Parcours introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_formateur_create_lecon(request, parcours_id):
    try:
        from iot.models import Parcours, Lecon
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        parcours = Parcours.objects.get(id=parcours_id, organisation=fp.organisation)
        titre = request.data.get("titre", "")
        ordre = request.data.get("ordre", 0)
        resume = request.data.get("resume", "")
        if not titre:
            return Response({"error": "Titre requis"}, status=400)
        l = Lecon.objects.create(parcours=parcours, titre=titre, ordre=ordre, resume=resume)
        return Response({"id": l.id, "titre": l.titre, "ordre": l.ordre}, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_formateur_create_bloc(request, lecon_id):
    try:
        from iot.models import Lecon, BlocPedagogique
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        lecon = Lecon.objects.get(id=lecon_id, parcours__organisation=fp.organisation)
        type_ = request.data.get("type", "texte")
        contenu = request.data.get("contenu", "")
        code = request.data.get("code", "")
        language = request.data.get("language", "")
        ordre = request.data.get("ordre", 0)
        b = BlocPedagogique.objects.create(lecon=lecon, type=type_, contenu=contenu, code=code, language=language, ordre=ordre)
        media_file = request.FILES.get("media_file")
        if media_file:
            b.media_file = media_file
            b.save(update_fields=["media_file"])
        return Response({
            "id": b.id, "type": b.type, "ordre": b.ordre,
            "media_url": request.build_absolute_uri(b.media_file.url) if b.media_file else None,
        }, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_formateur_create_quiz(request, lecon_id):
    try:
        from iot.models import Lecon, Quiz
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        lecon = Lecon.objects.get(id=lecon_id, parcours__organisation=fp.organisation)
        question = request.data.get("question", "")
        choix_a = request.data.get("choix_a", "")
        choix_b = request.data.get("choix_b", "")
        choix_c = request.data.get("choix_c", "")
        choix_d = request.data.get("choix_d", "")
        bonne_reponse = request.data.get("bonne_reponse", "A").upper()
        explication = request.data.get("explication", "")
        if not question or not choix_a or not choix_b:
            return Response({"error": "Question, choix A et B requis"}, status=400)
        q = Quiz.objects.create(
            lecon=lecon, question=question,
            choix_a=choix_a, choix_b=choix_b,
            choix_c=choix_c or None, choix_d=choix_d or None,
            bonne_reponse=bonne_reponse, explication=explication,
        )
        return Response({"id": q.id, "question": q.question}, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_formateur_lecon_detail(request, lecon_id):
    try:
        from iot.models import Lecon, BlocPedagogique, Quiz
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        lecon = Lecon.objects.get(id=lecon_id, parcours__organisation=fp.organisation)
        if request.method == "DELETE":
            lecon.delete()
            return Response({"deleted": True})
        if request.method == "PATCH":
            for field in ["titre", "ordre", "resume"]:
                if field in request.data:
                    setattr(lecon, field, request.data[field])
            lecon.save()
        blocs = [{
            "id": b.id, "type": b.type, "contenu": b.contenu, "code": b.code,
            "language": b.language, "ordre": b.ordre,
            "media_url": request.build_absolute_uri(b.media_file.url) if b.media_file else None,
        } for b in BlocPedagogique.objects.filter(lecon=lecon).order_by("ordre")]
        quiz_qs = Quiz.objects.filter(lecon=lecon).order_by("id")
        quiz = [{"id": q.id, "question": q.question, "choix_a": q.choix_a, "choix_b": q.choix_b,
                 "choix_c": q.choix_c, "choix_d": q.choix_d, "bonne_reponse": q.bonne_reponse,
                 "explication": q.explication} for q in quiz_qs]
        return Response({"id": lecon.id, "titre": lecon.titre, "ordre": lecon.ordre,
                         "resume": lecon.resume, "blocs": blocs, "quiz": quiz})
    except Lecon.DoesNotExist:
        return Response({"error": "Leçon introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_formateur_delete_bloc(request, bloc_id):
    try:
        from iot.models import BlocPedagogique
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        bloc = BlocPedagogique.objects.get(id=bloc_id, lecon__parcours__organisation=fp.organisation)
        if request.method == "DELETE":
            bloc.delete()
            return Response({"deleted": True})
        for field in ["type", "contenu", "code", "language", "ordre"]:
            if field in request.data:
                setattr(bloc, field, request.data[field])
        media_file = request.FILES.get("media_file")
        if media_file:
            bloc.media_file = media_file
        bloc.save()
        return Response({
            "id": bloc.id, "type": bloc.type, "ordre": bloc.ordre,
            "media_url": request.build_absolute_uri(bloc.media_file.url) if bloc.media_file else None,
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_formateur_delete_quiz(request, quiz_id):
    try:
        from iot.models import Quiz
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        q = Quiz.objects.get(id=quiz_id, lecon__parcours__organisation=fp.organisation)
        if request.method == "DELETE":
            q.delete()
            return Response({"deleted": True})
        for field in ["question", "choix_a", "choix_b", "choix_c", "choix_d", "bonne_reponse", "explication"]:
            if field in request.data:
                setattr(q, field, request.data[field])
        q.save()
        return Response({"id": q.id, "question": q.question})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def mobile_formateur_projects(request, parcours_id):
    try:
        from iot.models import Parcours, Project
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        parcours = Parcours.objects.get(id=parcours_id, organisation=fp.organisation)

        if request.method == "POST":
            titre = request.data.get("titre", "")
            description = request.data.get("description", "")
            if not titre:
                return Response({"error": "Titre requis"}, status=400)
            p = Project.objects.create(
                parcours=parcours, titre=titre, description=description,
                ordre=request.data.get("ordre", 0),
                language=request.data.get("language", ""),
                code=request.data.get("code", ""),
            )
            image = request.FILES.get("image")
            video = request.FILES.get("video")
            update_fields = []
            if image:
                p.image = image; update_fields.append("image")
            if video:
                p.video = video; update_fields.append("video")
            if update_fields:
                p.save(update_fields=update_fields)
            return Response({"id": p.id, "titre": p.titre}, status=201)

        projects = Project.objects.filter(parcours=parcours).order_by("ordre")
        return Response([{
            "id": p.id, "titre": p.titre, "description": p.description,
            "ordre": p.ordre, "language": p.language, "code": p.code,
            "image": request.build_absolute_uri(p.image.url) if p.image else None,
            "video": request.build_absolute_uri(p.video.url) if p.video else None,
        } for p in projects])
    except Parcours.DoesNotExist:
        return Response({"error": "Parcours introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_formateur_delete_project(request, project_id):
    try:
        from iot.models import Project
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        p = Project.objects.get(id=project_id, parcours__organisation=fp.organisation)
        if request.method == "DELETE":
            p.delete()
            return Response({"deleted": True})
        for field in ["titre", "description", "ordre", "language", "code"]:
            if field in request.data:
                setattr(p, field, request.data[field])
        image = request.FILES.get("image")
        video = request.FILES.get("video")
        if image:
            p.image = image
        if video:
            p.video = video
        p.save()
        return Response({
            "id": p.id, "titre": p.titre,
            "image": request.build_absolute_uri(p.image.url) if p.image else None,
            "video": request.build_absolute_uri(p.video.url) if p.video else None,
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_formateur_progressions(request):
    try:
        from iot.models import Progression, Lecon
        fp = _get_formateur(request.user)
        if not fp:
            return Response({"error": "Profil formateur requis"}, status=403)
        progs = Progression.objects.filter(
            lecon__parcours__organisation=fp.organisation
        ).select_related("user", "lecon", "lecon__parcours").order_by("-updated_at")[:50]
        return Response([{
            "apprenant": p.user.username,
            "parcours": p.lecon.parcours.titre,
            "lecon": p.lecon.titre,
            "completed": p.completed,
            "score": p.score,
            "updated_at": p.updated_at.isoformat(),
        } for p in progs])
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# MODULE EDUCATION (iot)
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_education_parcours(request):
    try:
        from iot.models import Parcours, Lecon, Progression, Certification
        parcours_qs = Parcours.objects.select_related("organisation").filter(is_published=True)
        data = []
        for p in parcours_qs:
            lecons = Lecon.objects.filter(parcours=p)
            total = lecons.count()
            completed = Progression.objects.filter(user=request.user, lecon__in=lecons, completed=True).count()
            pct = int((completed / total) * 100) if total else 0
            cert = Certification.objects.filter(user=request.user, parcours=p).first()
            data.append({
                "id": p.id,
                "titre": p.titre,
                "slug": p.slug,
                "description": p.description,
                "niveau": p.niveau,
                "certifiant": p.certifiant,
                "organisation": p.organisation.nom if p.organisation else "",
                "total_lecons": total,
                "completed_lecons": completed,
                "pourcentage": pct,
                "termine": pct == 100,
                "certifie": cert is not None,
                "score_final": cert.score_final if cert else None,
            })
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_education_lecons(request, parcours_id):
    try:
        from iot.models import Lecon, BlocPedagogique, Progression, Quiz
        lecons = Lecon.objects.filter(parcours_id=parcours_id).order_by("ordre")
        data = []
        for l in lecons:
            blocs = []
            for b in BlocPedagogique.objects.filter(lecon=l).order_by("ordre"):
                blocs.append({
                    "id": b.id,
                    "type": b.type,
                    "contenu": b.contenu,
                    "code": b.code,
                    "language": b.language,
                    "ordre": b.ordre,
                    "media_url": request.build_absolute_uri(b.media_file.url) if b.media_file else None,
                })
            prog = Progression.objects.filter(user=request.user, lecon=l).first()
            quiz_count = Quiz.objects.filter(lecon=l).count()
            # Determine if this leçon is locked
            if l.ordre == 1:
                is_locked = False
            else:
                precedente = Lecon.objects.filter(parcours=l.parcours, ordre__lt=l.ordre).order_by('-ordre').first()
                if not precedente:
                    is_locked = False
                else:
                    prog_prec = Progression.objects.filter(user=request.user, lecon=precedente).first()
                    is_locked = not (prog_prec and prog_prec.completed)
            data.append({
                "id": l.id,
                "titre": l.titre,
                "resume": l.resume,
                "ordre": l.ordre,
                "blocs": blocs,
                "quiz_count": quiz_count,
                "completed": prog.completed if prog else False,
                "score": prog.score if prog else 0,
                "is_locked": is_locked,
            })
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_education_quiz(request, lecon_id):
    try:
        from iot.models import Quiz
        quizzes = Quiz.objects.filter(lecon_id=lecon_id)
        data = []
        for q in quizzes:
            choices = {"A": q.choix_a, "B": q.choix_b}
            if q.choix_c:
                choices["C"] = q.choix_c
            if q.choix_d:
                choices["D"] = q.choix_d
            data.append({
                "id": q.id,
                "question": q.question,
                "choices": choices,
            })
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_education_submit_quiz(request, lecon_id):
    """
    Body: { "answers": {"quiz_id": "A", ...} }
    Returns: { "score": 80, "total": 5, "correct": 4, "details": [...] }
    """
    try:
        from iot.models import Quiz, Lecon, Progression
        answers = request.data.get("answers", {})
        quizzes = Quiz.objects.filter(lecon_id=lecon_id)
        total = quizzes.count()
        if total == 0:
            return Response({"error": "Aucun quiz pour cette leçon"}, status=400)
        correct = 0
        details = []
        for q in quizzes:
            user_ans = answers.get(str(q.id), "")
            is_correct = user_ans.upper() == q.bonne_reponse
            if is_correct:
                correct += 1
            details.append({
                "id": q.id,
                "question": q.question,
                "your_answer": user_ans,
                "correct_answer": q.bonne_reponse,
                "is_correct": is_correct,
                "explication": q.explication,
            })
        score = round((correct / total) * 100, 1)
        lecon = Lecon.objects.get(id=lecon_id)
        prog, _ = Progression.objects.get_or_create(user=request.user, lecon=lecon)
        prog.score = score
        prog.completed = score >= 50
        prog.save()

        # Check if all leçons in the parcours are now completed → generate certificate
        certificat_genere = False
        certificat_info = None
        if prog.completed:
            from iot.models import Parcours, Certification
            parcours = lecon.parcours
            all_lecons = list(Lecon.objects.filter(parcours=parcours))
            all_completed = all(
                Progression.objects.filter(user=request.user, lecon=l, completed=True).exists()
                for l in all_lecons
            )
            if all_completed and len(all_lecons) > 0:
                scores = Progression.objects.filter(user=request.user, lecon__parcours=parcours)
                score_final = round(sum(p.score for p in scores) / scores.count(), 1) if scores.count() > 0 else score
                cert, created = Certification.objects.get_or_create(
                    user=request.user,
                    parcours=parcours,
                    defaults={"score_final": score_final},
                )
                certificat_genere = created
                certificat_info = {"parcours": parcours.titre, "score_final": cert.score_final}

        return Response({
            "score": score, "total": total, "correct": correct,
            "completed": prog.completed, "details": details,
            "certificat_genere": certificat_genere,
            "certificat": certificat_info,
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_education_projects(request, parcours_id):
    try:
        from iot.models import Project
        projects = Project.objects.filter(parcours_id=parcours_id).order_by("ordre")
        data = [{
            "id": p.id,
            "titre": p.titre,
            "description": p.description,
            "language": p.language,
            "code": p.code,
            "ordre": p.ordre,
            "image": request.build_absolute_uri(p.image.url) if p.image else None,
            "video": request.build_absolute_uri(p.video.url) if p.video else None,
        } for p in projects]
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_education_certificats(request):
    try:
        from iot.models import Certification
        certs = Certification.objects.filter(user=request.user).select_related("parcours")
        data = [{
            "id": str(c.uuid),
            "parcours": c.parcours.titre if c.parcours else "",
            "score_final": c.score_final,
            "is_valid": c.is_valid,
            "created_at": c.created_at.isoformat(),
            "pdf_url": request.build_absolute_uri(c.pdf.url) if c.pdf else None,
            "qr_code_url": request.build_absolute_uri(c.qr_code.url) if c.qr_code else None,
            "verify_url": f"https://founatek224.pythonanywhere.com/certificat/{c.uuid}/",
        } for c in certs]
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# MODULE MONETISATION
# ══════════════════════════════════════════════════════════════════════════════

PLAN_PRICES = {"free": 0, "basic": 0.1, "pro": 15}

@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_dashboard(request):
    try:
        from monetisation.models import Wallet, Subscription, Transaction, PaymentRequest
        from monetisation.utils import PLAN_LIMITS, check_api_quota, check_device_quota
        wallet = Wallet.objects.filter(user=request.user).first()
        sub = Subscription.objects.filter(user=request.user).first()
        txns = Transaction.objects.filter(user=request.user).order_by("-timestamp")[:20]
        payments = PaymentRequest.objects.filter(user=request.user).order_by("-created_at")[:10]

        plan = sub.plan if sub else "free"
        limits = PLAN_LIMITS.get(plan, PLAN_LIMITS["free"])
        _, api_calls_used, max_calls = check_api_quota(request.user)
        _, device_count, max_devices = check_device_quota(request.user)

        return Response({
            "balance": wallet.balance if wallet else 0,
            "subscription": {
                "plan": plan,
                "start_date": sub.start_date.isoformat() if sub else None,
                "end_date": sub.end_date.isoformat() if sub and sub.end_date else None,
            },
            "usage": {
                "api_calls": api_calls_used,
                "max_calls": max_calls,
                "device_count": device_count,
                "max_devices": max_devices,
                "history_days": limits["history_days"],
                "alerts": limits["alerts"],
                "marketplace": limits["marketplace"],
                "pdf_reports": limits["pdf_reports"],
            },
            "plans": {
                key: {**vals, "price": PLAN_PRICES.get(key, 0)}
                for key, vals in PLAN_LIMITS.items()
            },
            "transactions": [{
                "id": t.id,
                "type": t.type,
                "amount": t.amount,
                "description": t.description,
                "timestamp": t.timestamp.isoformat(),
            } for t in txns],
            "payments": [{
                "id": p.id,
                "provider": p.provider,
                "amount": p.amount,
                "status": p.status,
                "created_at": p.created_at.isoformat(),
            } for p in payments],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_upgrade(request):
    """Body: { "plan": "basic" | "pro" | "free" }"""
    try:
        from monetisation.models import Wallet, Subscription, Transaction
        from django.utils import timezone
        from datetime import timedelta
        plan = request.data.get("plan", "free")
        if plan not in PLAN_PRICES:
            return Response({"error": "Plan invalide"}, status=400)
        price = PLAN_PRICES[plan]
        if price > 0:
            wallet, _ = Wallet.objects.get_or_create(user=request.user, defaults={"balance": 0})
            if wallet.balance < price:
                return Response({"error": "Solde insuffisant"}, status=402)
            wallet.balance -= price
            wallet.save()
            Transaction.objects.create(
                user=request.user, type="debit",
                amount=price, description=f"Abonnement {plan.upper()}"
            )
        sub, _ = Subscription.objects.get_or_create(user=request.user, defaults={"plan": "free"})
        sub.plan = plan
        sub.start_date = timezone.now().date()
        sub.end_date = (timezone.now() + timedelta(days=30)).date() if plan != "free" else None
        sub.save()
        return Response({"plan": sub.plan, "end_date": sub.end_date.isoformat() if sub.end_date else None})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_recharge(request):
    try:
        from monetisation.models import PaymentRequest
        provider = request.data.get("provider", "MTN")
        phone = request.data.get("phone_number", "")
        amount = float(request.data.get("amount", 0))
        transaction_id = request.data.get("transaction_id", "")
        if amount <= 0 or not phone:
            return Response({"error": "Paramètres invalides"}, status=400)
        pay = PaymentRequest.objects.create(
            user=request.user, provider=provider,
            phone_number=phone, amount=amount, status="pending",
            transaction_id=transaction_id or None,
        )
        return Response({"id": pay.id, "status": pay.status, "provider": pay.provider, "amount": pay.amount})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_merchant_numbers(request):
    """Numeros Mobile Money du marchand a qui envoyer le paiement (flux manuel)."""
    try:
        from monetisation.models import MobileMoneyAccount
        accounts = MobileMoneyAccount.objects.filter(is_active=True)
        return Response([{
            "id": a.id, "provider": a.provider, "phone_number": a.phone_number,
            "owner_name": a.owner_name, "instructions": a.instructions,
        } for a in accounts])
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_admin_pending(request):
    """Liste des demandes de paiement en attente (reserve au staff)."""
    if not request.user.is_staff:
        return Response({"error": "Accès réservé aux administrateurs"}, status=403)
    try:
        from monetisation.models import PaymentRequest
        pending = PaymentRequest.objects.filter(status="pending").order_by("-created_at")
        return Response([{
            "id": p.id, "username": p.user.username, "provider": p.provider,
            "phone_number": p.phone_number, "amount": p.amount,
            "transaction_id": p.transaction_id, "created_at": p.created_at.isoformat(),
        } for p in pending])
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_admin_review(request, payment_id):
    """Confirme ou refuse une demande de paiement (staff only). Body: {action: 'confirm'|'reject'}"""
    if not request.user.is_staff:
        return Response({"error": "Accès réservé aux administrateurs"}, status=403)
    try:
        from django.utils import timezone as _timezone
        from monetisation.models import PaymentRequest, Wallet, Transaction
        payment = PaymentRequest.objects.get(id=payment_id, status="pending")
        action = request.data.get("action")
        if action == "confirm":
            payment.status = "success"
            wallet, _ = Wallet.objects.get_or_create(user=payment.user, defaults={"balance": 0})
            wallet.balance += payment.amount
            wallet.save()
            Transaction.objects.create(
                user=payment.user, type="credit", amount=payment.amount,
                description=f"Recharge {payment.provider} confirmée par {request.user.username}",
            )
        elif action == "reject":
            payment.status = "failed"
        else:
            return Response({"error": "action doit être 'confirm' ou 'reject'"}, status=400)
        payment.reviewed_by = request.user
        payment.reviewed_at = _timezone.now()
        payment.save()
        return Response({"id": payment.id, "status": payment.status})
    except PaymentRequest.DoesNotExist:
        return Response({"error": "Demande introuvable ou déjà traitée"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_referral(request):
    try:
        from monetisation.models import Referral, ReferralTransaction
        referral = Referral.objects.filter(referrer=request.user).first()
        if not referral:
            import uuid as uuid_lib
            referral = Referral.objects.create(referrer=request.user, referred=request.user, code=str(uuid_lib.uuid4())[:8])
        referred_users = Referral.objects.filter(referrer=request.user).exclude(referred=request.user)
        commissions = ReferralTransaction.objects.filter(referral__referrer=request.user).order_by("-created_at")[:20]
        total_points = sum(c.amount for c in commissions if not c.converted)
        return Response({
            "code": referral.code,
            "referred_count": referred_users.count(),
            "total_points": total_points,
            "commissions": [{
                "id": c.id,
                "amount": c.amount,
                "description": c.description,
                "converted": c.converted,
                "created_at": c.created_at.isoformat(),
            } for c in commissions],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_convert(request):
    """Convertit les points de parrainage en crédits wallet (1000 pts = 1 EUR)"""
    try:
        from monetisation.models import Referral, ReferralTransaction, Wallet, Transaction
        commissions = ReferralTransaction.objects.filter(
            referral__referrer=request.user, converted=False
        )
        total = sum(c.amount for c in commissions)
        if total < 1000:
            return Response({"error": "Minimum 1000 points requis"}, status=400)
        euros = round(total / 1000, 2)
        commissions.update(converted=True)
        wallet, _ = Wallet.objects.get_or_create(user=request.user, defaults={"balance": 0})
        wallet.balance += euros
        wallet.save()
        Transaction.objects.create(
            user=request.user, type="credit",
            amount=euros, description=f"Conversion {int(total)} points de parrainage"
        )
        return Response({"converted_points": total, "euros_credited": euros, "new_balance": wallet.balance})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_transactions(request):
    """Historique complet des transactions wallet (pagine par page= et page_size=)."""
    try:
        from monetisation.models import Transaction
        page = max(1, int(request.query_params.get("page", 1)))
        page_size = min(100, max(1, int(request.query_params.get("page_size", 30))))
        qs = Transaction.objects.filter(user=request.user).order_by("-timestamp")
        total = qs.count()
        start = (page - 1) * page_size
        items = qs[start:start + page_size]
        return Response({
            "count": total,
            "page": page,
            "page_size": page_size,
            "has_more": start + page_size < total,
            "results": [{
                "id": t.id, "type": t.type, "amount": t.amount,
                "description": t.description, "timestamp": t.timestamp.isoformat(),
            } for t in items],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_paypal_create_order(request):
    """Cree un ordre PayPal reel. Body: {amount, type: 'credit'|'subscription', plan?}"""
    try:
        import requests as _requests
        from django.conf import settings as _settings
        amount = request.data.get("amount")
        payment_type = request.data.get("type", "credit")
        plan = request.data.get("plan")
        if not amount:
            return Response({"error": "Montant requis"}, status=400)
        url = f"{_settings.PAYPAL_API_BASE}/v2/checkout/orders"
        auth = (_settings.PAYPAL_CLIENT_ID, _settings.PAYPAL_SECRET)
        payload = {
            "intent": "CAPTURE",
            "purchase_units": [{
                "amount": {"currency_code": "EUR", "value": str(amount)},
                "custom_id": f"{payment_type}|{plan or ''}",
            }],
        }
        resp = _requests.post(url, json=payload, auth=auth, timeout=10)
        if resp.status_code == 201:
            return Response(resp.json(), status=201)
        return Response({"error": "Impossible de creer la commande PayPal"}, status=400)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_monetisation_paypal_capture_order(request):
    """Capture un ordre PayPal et credite le wallet / active l'abonnement. Body: {orderID}"""
    try:
        import requests as _requests
        from django.conf import settings as _settings
        from django.utils import timezone as _timezone
        from monetisation.models import Wallet, Subscription, Transaction, PaymentRequest, ReferralTransaction

        order_id = request.data.get("orderID")
        if not order_id:
            return Response({"error": "orderID requis"}, status=400)

        url = f"{_settings.PAYPAL_API_BASE}/v2/checkout/orders/{order_id}/capture"
        auth = (_settings.PAYPAL_CLIENT_ID, _settings.PAYPAL_SECRET)
        resp = _requests.post(url, auth=auth, timeout=10)
        result = resp.json()

        if result.get("status") != "COMPLETED":
            return Response({"success": False, "error": "Paiement non complete"}, status=400)

        purchase_unit = result["purchase_units"][0]
        custom_id = purchase_unit.get("custom_id", "")
        amount = float(purchase_unit["payments"]["captures"][0]["amount"]["value"])
        payment_type, plan = custom_id.split("|") if "|" in custom_id else ("credit", "")

        wallet, _ = Wallet.objects.get_or_create(user=request.user, defaults={"balance": 0})

        if payment_type == "credit":
            wallet.balance += amount
            wallet.save()
            Transaction.objects.create(user=request.user, type="credit", amount=amount, description="Recharge via PayPal")
        elif payment_type == "subscription" and plan:
            sub, _ = Subscription.objects.get_or_create(user=request.user, defaults={"plan": "free"})
            sub.plan = plan
            sub.end_date = None if plan == "free" else _timezone.now().date().replace(year=_timezone.now().year + 1)
            sub.save()
            Transaction.objects.create(user=request.user, type="subscription", amount=amount, description=f"Abonnement {plan} via PayPal")

        if hasattr(request.user, "referred_by"):
            ref = request.user.referred_by
            commission_points = amount * 0.05 * 1000
            ReferralTransaction.objects.create(
                referral=ref, amount=commission_points,
                description=f"Commission sur {payment_type} de {request.user.username}", converted=False,
            )

        try:
            payment = PaymentRequest.objects.get(user=request.user, amount=amount, status="pending", provider="paypal")
            payment.status = "success"
            payment.save()
        except PaymentRequest.DoesNotExist:
            pass

        return Response({"success": True, "new_balance": wallet.balance})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# MODULE TRANSPARENCE PRODUIT
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def mobile_transparence_company(request):
    try:
        from product_transparency.models import Company
        company = Company.objects.filter(user=request.user).first()
        if not company:
            return Response({"error": "Aucune entreprise associée à votre compte"}, status=404)
        if request.method == "PATCH":
            company.name = request.data.get("name", company.name)
            company.currency = request.data.get("currency", company.currency)
            update_fields = ["name", "currency"]
            logo = request.FILES.get("logo")
            if logo:
                company.logo = logo
                update_fields.append("logo")
            company.save(update_fields=update_fields)
        return Response({
            "id": company.id,
            "name": company.name,
            "slug": company.slug,
            "currency": company.currency,
            "is_active": company.is_active,
            "created_at": company.created_at.isoformat(),
            "logo": request.build_absolute_uri(company.logo.url) if company.logo else None,
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_transparence_products(request):
    try:
        from product_transparency.models import Product, Company
        from django.utils.timezone import now
        company = Company.objects.filter(user=request.user).first()
        if company is None:
            return Response([])
        products = Product.objects.filter(company=company).prefetch_related("pricing")
        data = []
        for p in products:
            pricing = getattr(p, "pricing", None)
            qr = getattr(p, "qr", None)
            qr_url = f"https://founatek224.pythonanywhere.com/product_transparency/dashboard/products/{p.id}/qr/download/" if qr else None
            status_label = None
            if pricing:
                status_label = "EXPIRÉ" if pricing.expiry_date < now().date() else "VALIDE"
            data.append({
                "id": p.id,
                "name": p.name,
                "sku": p.sku,
                "uuid": str(p.uuid),
                "price": float(pricing.price) if pricing else None,
                "expiry_date": pricing.expiry_date.isoformat() if pricing else None,
                "production_date": pricing.production_date.isoformat() if pricing else None,
                "status": status_label,
                "image": request.build_absolute_uri(p.image.url) if p.image else None,
                "qr_download_url": qr_url,
            })
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_transparence_create_product(request):
    try:
        from product_transparency.models import Company, Product, ProductPricing
        from decimal import Decimal
        company = Company.objects.filter(user=request.user).first()
        if not company:
            return Response({"error": "Aucune entreprise associée"}, status=400)
        name = request.data.get("name", "")
        sku = request.data.get("sku", "")
        if not name or not sku:
            return Response({"error": "Nom et SKU requis"}, status=400)
        if Product.objects.filter(sku=sku).exists():
            return Response({"error": "Ce SKU existe déjà"}, status=409)
        product = Product.objects.create(company=company, name=name, sku=sku)
        image = request.FILES.get("image")
        if image:
            product.image = image
            product.save(update_fields=["image"])
        price = request.data.get("price")
        prod_date = request.data.get("production_date")
        exp_date = request.data.get("expiry_date")
        if price and prod_date and exp_date:
            ProductPricing.objects.create(
                product=product,
                price=Decimal(str(price)),
                production_date=prod_date,
                expiry_date=exp_date,
            )
        return Response({"id": product.id, "uuid": str(product.uuid), "name": product.name, "sku": product.sku}, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_transparence_update_pricing(request, product_id):
    try:
        from product_transparency.models import Company, Product, ProductPricing, ProductPriceHistory
        from decimal import Decimal
        company = Company.objects.filter(user=request.user).first()
        product = Product.objects.get(id=product_id, company=company)
        price = Decimal(str(request.data.get("price", 0)))
        prod_date = request.data.get("production_date")
        exp_date = request.data.get("expiry_date")
        reason = request.data.get("reason", "")
        pricing, _ = ProductPricing.objects.get_or_create(
            product=product,
            defaults={"price": price, "production_date": prod_date, "expiry_date": exp_date}
        )
        if not _:
            ProductPriceHistory.objects.create(
                product=product, price=pricing.price,
                production_date=pricing.production_date,
                expiry_date=pricing.expiry_date,
                changed_by=request.user, reason=reason,
            )
            pricing.price = price
            pricing.production_date = prod_date
            pricing.expiry_date = exp_date
            pricing.save()
        return Response({"price": float(pricing.price), "expiry_date": str(pricing.expiry_date)})
    except Product.DoesNotExist:
        return Response({"error": "Produit introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_transparence_update_product_image(request, product_id):
    try:
        from product_transparency.models import Company, Product
        company = Company.objects.filter(user=request.user).first()
        product = Product.objects.get(id=product_id, company=company)
        image = request.FILES.get("image")
        if not image:
            return Response({"error": "Aucune image envoyée"}, status=400)
        product.image = image
        product.save(update_fields=["image"])
        return Response({"image": request.build_absolute_uri(product.image.url)})
    except Product.DoesNotExist:
        return Response({"error": "Produit introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def mobile_transparence_update_product(request, product_id):
    """Edition complete (nom/SKU + image optionnelle) d'un produit existant."""
    try:
        from product_transparency.models import Company, Product
        company = Company.objects.filter(user=request.user).first()
        product = Product.objects.get(id=product_id, company=company)
        name = request.data.get("name")
        sku = request.data.get("sku")
        if sku and Product.objects.filter(sku=sku).exclude(id=product.id).exists():
            return Response({"error": "Ce SKU existe déjà"}, status=409)
        if name:
            product.name = name
        if sku:
            product.sku = sku
        image = request.FILES.get("image")
        if image:
            product.image = image
        product.save()
        return Response({
            "id": product.id, "name": product.name, "sku": product.sku,
            "image": request.build_absolute_uri(product.image.url) if product.image else None,
        })
    except Product.DoesNotExist:
        return Response({"error": "Produit introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_transparence_download_qr(request, product_id):
    """Telecharge le QR code PNG d'un produit."""
    try:
        from django.http import FileResponse, Http404
        from product_transparency.models import Company, Product
        company = Company.objects.filter(user=request.user).first()
        product = Product.objects.filter(id=product_id, company=company).select_related("qr").first()
        if not product or not hasattr(product, "qr") or not product.qr.qr_code:
            raise Http404("QR code introuvable")
        return FileResponse(product.qr.qr_code.open("rb"), as_attachment=True, filename=f"{product.sku}_QR.png")
    except Http404:
        return Response({"error": "QR code introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_transparence_qr_labels_pdf(request):
    """Genere un PDF de planches d'etiquettes QR pour tous les produits de l'entreprise."""
    try:
        from django.http import FileResponse
        from product_transparency.models import Company, Product
        from product_transparency.services.qr_label_pdf import generate_qr_labels_pdf
        company = Company.objects.filter(user=request.user).first()
        if not company:
            return Response({"error": "Aucune entreprise associée"}, status=400)
        per_page = int(request.query_params.get("per_page", 12))
        products = list(Product.objects.filter(company=company, qr__isnull=False).select_related("company", "qr"))
        if not products:
            return Response({"error": "Aucun produit avec QR code"}, status=400)
        pdf_buffer = generate_qr_labels_pdf(products, per_page)
        return FileResponse(pdf_buffer, as_attachment=True, filename="qr_labels.pdf", content_type="application/pdf")
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_transparence_scan(request, uuid):
    try:
        from product_transparency.models import Product, ProductPriceHistory
        from django.utils.timezone import now
        p = Product.objects.select_related("company").get(uuid=uuid)
        pricing = getattr(p, "pricing", None)
        history = ProductPriceHistory.objects.filter(product=p).order_by("-changed_at")[:10]
        status_label = None
        if pricing:
            status_label = "EXPIRÉ" if pricing.expiry_date < now().date() else "VALIDE"
        return Response({
            "id": p.id,
            "name": p.name,
            "sku": p.sku,
            "company": p.company.name if p.company else "",
            "currency": p.company.currency if p.company else "GNF",
            "price": float(pricing.price) if pricing else None,
            "expiry_date": pricing.expiry_date.isoformat() if pricing else None,
            "production_date": pricing.production_date.isoformat() if pricing else None,
            "status": status_label,
            "image": request.build_absolute_uri(p.image.url) if p.image else None,
            "company_logo": request.build_absolute_uri(p.company.logo.url) if p.company and p.company.logo else None,
            "price_history": [{
                "price": float(h.price),
                "changed_at": h.changed_at.isoformat(),
                "reason": h.reason,
            } for h in history],
        })
    except Product.DoesNotExist:
        return Response({"error": "Produit introuvable"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_transparence_create_sale(request):
    """Body: { "items": [{"product_id": 1, "quantity": 2, "unit_price": 5000}] }"""
    try:
        from product_transparency.models import Company, Product, Sale, SaleItem
        from decimal import Decimal
        company = Company.objects.filter(user=request.user).first()
        if not company:
            return Response({"error": "Aucune entreprise associée"}, status=400)
        items_data = request.data.get("items", [])
        if not items_data:
            return Response({"error": "Aucun article"}, status=400)
        total = sum(Decimal(str(i["unit_price"])) * int(i["quantity"]) for i in items_data)
        sale = Sale.objects.create(company=company, total_amount=total)
        for i in items_data:
            product = Product.objects.get(id=i["product_id"], company=company)
            SaleItem.objects.create(
                sale=sale, product=product,
                quantity=int(i["quantity"]),
                unit_price=Decimal(str(i["unit_price"])),
            )
        return Response({"id": str(sale.id), "total": float(sale.total_amount), "items": len(items_data)}, status=201)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def mobile_relay_delete(request, num):
    try:
        relay = Relais.objects.get(num=num, user=request.user)
        relay.delete()
        return Response({"deleted": True})
    except Relais.DoesNotExist:
        return _json_error("Relais introuvable", 404)


@csrf_exempt
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def mobile_relay_rename(request, num):
    try:
        relay = Relais.objects.get(num=num, user=request.user)
    except Relais.DoesNotExist:
        return _json_error("Relais introuvable", 404)
    nom = request.data.get("nom", relay.nom)
    relay.nom = nom
    relay.save(update_fields=["nom"])
    return Response({"num": relay.num, "nom": relay.nom})


# ── Comptage liste + reset ─────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET", "DELETE"])
@permission_classes([IsAuthenticated])
def mobile_comptage_list(request):
    if request.method == "DELETE":
        Comptage.objects.filter(user=request.user).delete()
        return Response({"reset": True})
    items = Comptage.objects.filter(user=request.user).order_by("-timestamp")[:20]
    data = [{"id": c.id, "compteur": c.compteur, "timestamp": c.timestamp.isoformat()} for c in items]
    return Response(data)


# ── Access logs complet ────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_access_logs_full(request):
    from .models import AccessLog
    logs = AccessLog.objects.order_by("-timestamp")[:50]
    data = []
    for log in logs:
        data.append({
            "id": log.id,
            "badge_id": log.badge_id if hasattr(log, "badge_id") else str(log.badge) if hasattr(log, "badge") else "",
            "access_granted": log.access_granted if hasattr(log, "access_granted") else log.granted if hasattr(log, "granted") else None,
            "timestamp": log.timestamp.isoformat() if hasattr(log, "timestamp") else "",
            "door": str(log.door) if hasattr(log, "door") else "",
        })
    return Response(data)


# ── Profil utilisateur ─────────────────────────────────────────────────────────
@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_profile(request):
    user = request.user
    try:
        token_obj = Token.objects.get(user=user)
        token_key = token_obj.key
    except Token.DoesNotExist:
        token_key = None

    avatar_url = None
    try:
        from espcontrol.models import UserProfile
        up = UserProfile.objects.get(user=user)
        if up.avatar:
            avatar_url = request.build_absolute_uri(up.avatar.url)
    except Exception:
        pass

    return Response({
        "username": user.username,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "is_staff": user.is_staff,
        "date_joined": user.date_joined.isoformat(),
        "token": token_key,
        "avatar_url": avatar_url,
        "counts": {
            "relays":   Relais.objects.filter(user=user).count(),
            "dht":      DHTData.objects.filter(user=user).count(),
            "gas":      SensorData.objects.filter(user=user).count(),
            "ntc":      NtcSensorData.objects.filter(user=user).count(),
            "soil":     SoilData.objects.filter(user=user).count(),
            "devices":  Device.objects.filter(user=user).count(),
            "alerts":   AgentAlert.objects.filter(user=user).count(),
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
# MODULE QUALITE DE L'AIR — dashboard multi-villes, prevision, alertes
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_air_map(request):
    """Carte multi-villes (Conakry, Dakar, Abidjan, Lagos, Nairobi)."""
    try:
        from espcontrol.models import AirReading
        cities = ["conakry", "dakar", "abidjan", "lagos", "nairobi"]
        data = []
        for city in cities:
            reading = AirReading.objects.filter(source__icontains=city.upper(), is_active=True).order_by("-simulated_time").first()
            if not reading:
                continue
            data.append({
                "city": city, "lat": reading.latitude, "lon": reading.longitude,
                "pm2p5": reading.pm2p5, "pm10": reading.pm10, "co": reading.co, "no2": reading.no2,
                "confidence": reading.confidence, "origin": reading.origin, "source": reading.source,
            })
        return Response({"data": data})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_air_city_latest(request, city):
    """Derniere lecture (reelle/virtuelle/hybride) pour une ville donnee."""
    try:
        from espcontrol.ml.services.air_fusion import get_latest_air_reading
        reading = get_latest_air_reading(city)
        if not reading:
            return Response({"error": "no data"}, status=404)
        return Response({
            "city": city, "origin": reading.origin, "confidence": reading.confidence,
            "simulated_time": reading.simulated_time.isoformat(), "created_at": reading.created_at.isoformat(),
            "location": {"lat": reading.latitude, "lon": reading.longitude},
            "data": {"pm2p5": reading.pm2p5, "pm10": reading.pm10, "co": reading.co, "no2": reading.no2},
            "source": reading.source,
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_air_forecast(request):
    """Prevision qualite de l'air a 3h (Conakry) + alertes associees."""
    try:
        from espcontrol.ml.services.air_predictor import predict_conakry_next_3h
        from espcontrol.ml.services.air_alerts import evaluate_air_quality
        prediction_data = predict_conakry_next_3h()
        alerts = evaluate_air_quality(prediction_data["prediction"])
        return Response({
            "from_time": prediction_data.get("from_time"),
            "to_time": prediction_data.get("to_time"),
            "prediction": prediction_data.get("prediction"),
            "global_status": alerts.get("global_status"),
            "alerts": alerts.get("alerts"),
        })
    except FileNotFoundError as e:
        return Response({"error": "Donnees ML manquantes", "detail": str(e)}, status=503)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_air_alert_now(request, city):
    """Alerte immediate basee sur la derniere mesure reelle/virtuelle d'une ville."""
    try:
        from espcontrol.models import AirReading
        from espcontrol.ml.services.air_alerts import evaluate_air_alert_now
        reading = AirReading.objects.filter(source__icontains=city.upper(), is_active=True).order_by("-simulated_time").first()
        if not reading:
            return Response({"alert": None})
        alert = evaluate_air_alert_now(reading)
        return Response({"city": city, "alert": alert})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# MODULE IRRIGATION ML
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mobile_irrigation_prediction(request):
    """Body: {temperature_c, humidity_air, rainfall_mm, hour} -> {irrigation: 'ON'|'OFF'}"""
    try:
        import pandas as pd
        import joblib
        from pathlib import Path
        from .MAL.features import FEATURES
    except Exception as e:
        return Response({"error": "Dépendances irrigation indisponibles", "detail": str(e)}, status=503)
    try:
        required = ["temperature_c", "humidity_air", "rainfall_mm", "hour"]
        missing = [k for k in required if k not in request.data]
        if missing:
            return Response({"error": "Champs manquants", "missing": missing}, status=400)

        model_path = Path(__file__).resolve().parent / "MAL" / "irrigation_model.pkl"
        if not model_path.exists():
            return Response({"error": "Modèle indisponible sur le serveur"}, status=503)
        model = joblib.load(model_path)

        sample_df = pd.DataFrame([{
            "temperature_c": request.data["temperature_c"],
            "humidity_air": request.data["humidity_air"],
            "rainfall_mm": request.data["rainfall_mm"],
            "hour": request.data["hour"],
        }], columns=FEATURES)
        prediction = model.predict(sample_df)[0]
        return Response({"irrigation": "ON" if int(prediction) == 1 else "OFF", "decision": int(prediction)})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTS EXCEL
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_export_data_excel(request):
    """Export multi-feuilles (DHT11, Gaz, NTC, Qualite de l'air) au format xlsx.
    Query params optionnels : start_date, end_date (AAAA-MM-JJ)."""
    try:
        import openpyxl
        from openpyxl.styles import Font
        from django.http import HttpResponse
        from .models import SoilData as _SoilData  # noqa
        from monetisation.models import Subscription

        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        user = request.user

        workbook = openpyxl.Workbook()
        header_font = Font(bold=True)

        # ── DHT11 (reserve aux plans basic/pro, comme sur le site) ──
        sub = Subscription.objects.filter(user=user).first()
        plan = sub.plan if sub else "free"
        dht_sheet = workbook.active
        dht_sheet.title = "DHT11"
        if plan in ("basic", "pro") or user.is_superuser:
            dht_data = DHTData.objects.filter(user=user)
            if start_date and end_date:
                dht_data = dht_data.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
            dht_sheet.append(["Date", "Température (°C)", "Humidité (%)"])
            for entry in dht_data.order_by("-created_at"):
                dht_sheet.append([entry.created_at.strftime("%Y-%m-%d %H:%M:%S"), entry.temperature, entry.humidity])
        else:
            dht_sheet.append(["Export DHT11 réservé aux plans Basic et Pro. Passez à un plan supérieur dans l'onglet Monétisation."])
        for cell in dht_sheet[1]:
            cell.font = header_font

        # ── Gaz ──
        gas_sheet = workbook.create_sheet("Gaz")
        gas_data = SensorData.objects.filter(user=user)
        if start_date and end_date:
            gas_data = gas_data.filter(timestamp__date__gte=start_date, timestamp__date__lte=end_date)
        gas_sheet.append(["Date", "Température (°C)", "Humidité (%)", "CO2 (ppm)", "Type de Gaz"])
        for entry in gas_data.order_by("-timestamp"):
            gas_sheet.append([entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"), entry.temperature, entry.humidity, entry.co2, entry.gaz_type])
        for cell in gas_sheet[1]:
            cell.font = header_font

        # ── NTC ──
        ntc_sheet = workbook.create_sheet("NTC")
        ntc_data = NtcSensorData.objects.filter(user=user)
        if start_date and end_date:
            ntc_data = ntc_data.filter(timestamp__date__gte=start_date, timestamp__date__lte=end_date)
        ntc_sheet.append(["Date et heure", "Température (°C)"])
        for entry in ntc_data.order_by("-timestamp"):
            ntc_sheet.append([entry.timestamp.strftime("%Y-%m-%d %H:%M:%S"), entry.temperature])
        for cell in ntc_sheet[1]:
            cell.font = header_font

        # ── Qualite de l'air (tous appareils, 1000 dernieres mesures) ──
        air_sheet = workbook.create_sheet("Qualité Air")
        devices = Device.objects.filter(user=user)
        air_data = AppareilData.objects.filter(device__in=devices).order_by("-received_at")[:1000]
        air_sheet.append(["Horodatage", "Appareil", "PM2.5 (µg/m³)", "PM10 (µg/m³)", "Température (°C)", "Humidité (%)", "Gaz (PPM)", "Latitude", "Longitude"])
        for entry in air_data:
            p = entry.payload or {}
            air_sheet.append([
                entry.received_at.strftime("%d/%m/%Y %H:%M:%S"), entry.device.name,
                p.get("pm2p5", "N/A"), p.get("pm10", "N/A"), p.get("temperature", "N/A"),
                p.get("humidity", "N/A"), p.get("mq135_ppm", "N/A"), p.get("latitude", "N/A"), p.get("longitude", "N/A"),
            ])
        for cell in air_sheet[1]:
            cell.font = header_font

        for sheet in workbook.worksheets:
            for col_cells in sheet.columns:
                sheet.column_dimensions[col_cells[0].column_letter].width = 22

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"Founatek_Export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# ARCHIVES DE SURVEILLANCE (photos avec detection de mouvement)
# ══════════════════════════════════════════════════════════════════════════════

@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_surveillance_archive(request):
    """Galerie paginee des photos avec detection de mouvement. Query: page, page_size."""
    try:
        page = max(1, int(request.query_params.get("page", 1)))
        page_size = min(50, max(1, int(request.query_params.get("page_size", 20))))
        qs = UploadedImage.objects.filter(user=request.user, has_motion=True).order_by("-created_at")
        total = qs.count()
        start = (page - 1) * page_size
        items = qs[start:start + page_size]
        return Response({
            "count": total, "page": page, "page_size": page_size,
            "has_more": start + page_size < total,
            "results": [{
                "id": img.id,
                "image_url": request.build_absolute_uri(img.image.url) if img.image else None,
                "created_at": img.created_at.isoformat(),
                "camera_id": getattr(img, "camera_id", None),
            } for img in items],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_surveillance_download_zip(request):
    """Telecharge un zip des 50 dernieres photos avec detection de mouvement."""
    try:
        import zipfile
        import os
        from django.http import HttpResponse
        images = UploadedImage.objects.filter(user=request.user, has_motion=True).order_by("-created_at")[:50]
        if not images:
            return Response({"error": "Aucune image"}, status=404)
        response = HttpResponse(content_type="application/zip")
        now_str = timezone.now().strftime("%Y%m%d_%H%M")
        response["Content-Disposition"] = f'attachment; filename="Preuves_Founatek_{now_str}.zip"'
        with zipfile.ZipFile(response, "w") as zip_file:
            for img in images:
                if img.image:
                    try:
                        zip_file.write(img.image.path, arcname=os.path.basename(img.image.path))
                    except Exception:
                        pass
        return response
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD "UNIVERS" MULTI-APPAREILS + ALERTES PAR CAPTEUR
# ══════════════════════════════════════════════════════════════════════════════

def _detect_anomalies_for_device(device, window=50, sigma_threshold=2):
    import numpy as _np
    data_points = list(AppareilData.objects.filter(device=device).order_by("-received_at")[:window])
    if not data_points:
        return []
    sensors = ["temperature", "humidity", "soil_moisture", "ultrason", "mq135_ppm", "pm2p5", "pm10"]
    anomalies = []
    for sensor in sensors:
        values = [d.payload.get(sensor) for d in data_points if d.payload and isinstance(d.payload.get(sensor), (int, float))]
        if len(values) < 5:
            continue
        mean = _np.mean(values)
        std = _np.std(values)
        latest_value = values[0]
        if std > 0 and abs(latest_value - mean) > sigma_threshold * std:
            anomalies.append({"sensor": sensor, "value": latest_value, "mean": round(float(mean), 2), "std": round(float(std), 2)})
    return anomalies


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_dashboard_univers(request):
    """Vue d'ensemble multi-appareils avec detection d'anomalies (Z-score) par capteur."""
    try:
        devices = Device.objects.filter(user=request.user)
        devices_data = []
        for device in devices:
            data_points = list(AppareilData.objects.filter(device=device).order_by("-received_at")[:50])
            anomalies = _detect_anomalies_for_device(device, window=50, sigma_threshold=2)
            devices_data.append({
                "device": {"id": device.id, "name": device.name, "device_id": device.device_id, "is_active": device.is_active},
                "data_points": [{
                    "id": p.id, "received_at": p.received_at.isoformat(), "payload": p.payload,
                } for p in reversed(data_points)],
                "has_anomaly": bool(anomalies),
                "anomalies": anomalies,
            })
        alerts = AgentAlert.objects.filter(user=request.user, is_read=False).order_by("-created_at")[:20]
        return Response({
            "devices": devices_data,
            "alerts": [{
                "id": a.id, "message": a.message, "level": a.level, "sensor": a.sensor,
                "created_at": a.created_at.isoformat(),
            } for a in alerts],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_alert_graph_data(request, sensor):
    """Donnees pour un graphique d'evolution des alertes d'un capteur donne."""
    try:
        alerts = AgentAlert.objects.filter(user=request.user, sensor=sensor).order_by("created_at")[:200]
        return Response({
            "labels": [a.created_at.strftime("%H:%M") for a in alerts],
            "values": [a.value for a in alerts],
            "levels": [a.level for a in alerts],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def mobile_alerts_by_sensor(request, sensor):
    """Historique des 100 dernieres alertes pour un capteur donne."""
    try:
        alerts = AgentAlert.objects.filter(user=request.user, sensor=sensor).order_by("-created_at")[:100]
        return Response([{
            "message": a.message, "value": a.value, "level": a.level,
            "created_at": a.created_at.isoformat(),
            "device": a.device.device_id if a.device else None,
        } for a in alerts])
    except Exception as e:
        return Response({"error": str(e)}, status=500)
